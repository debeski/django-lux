import io
import json
import zipfile
from datetime import datetime, timedelta
from pathlib import Path
from types import SimpleNamespace

from dlux.tests.harness import setup_test_environment

setup_test_environment()

import tempfile
from unittest.mock import patch

from django.apps import apps
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Permission
from django.core.files.storage import default_storage
from django.db import models
from django.test import Client, TestCase, override_settings
from django.urls import reverse
from django.utils import timezone

from dlux.reports import (
    backup_record_folder,
    build_model_entries_xlsx,
    build_report_chart_data,
    build_reports_overview,
    get_report_entry_fields,
    get_report_window_bounds,
    is_report_eligible_activity_model_name,
    is_report_eligible_model,
    run_report_backup,
    write_backup_zip,
)


def _load_workbook(content):
    from openpyxl import load_workbook
    return load_workbook(io.BytesIO(content))

User = get_user_model()

ACTIVITY_BACKUP_CONFIG = {
    'reports': {'include_models': ['dlux.activitylog']},
}


class NumberedBackupRecord(models.Model):
    number = models.CharField(max_length=100)

    class Meta:
        app_label = 'backup_tests'
        managed = False


class ConfiguredBackupRecord(models.Model):
    case_reference = models.CharField(max_length=100)

    class Meta:
        app_label = 'backup_tests'
        managed = False


def _make_logs():
    """One activity row inside the current week, one ~60 days old."""
    ActivityLog = apps.get_model('dlux', 'ActivityLog')
    recent = ActivityLog.objects.create(action='CREATE', model_name='Project Entry')
    old = ActivityLog.objects.create(action='CREATE', model_name='Project Entry')
    ActivityLog.objects.filter(pk=old.pk).update(
        created_at=timezone.now() - timedelta(days=60)
    )
    return recent, old


def _zip_activity_rows(zip_bytes):
    """Activity rows carried by the ZIP's entries workbook, plus its manifest.

    The reports ZIP ships data as the workbook (no serialized JSON), so row
    identity is read back from the sheet's ID column.
    """
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        workbook = _load_workbook(zf.read('entries.xlsx'))
        manifest = json.loads(zf.read('manifest.json'))
    sheet = workbook['Activity Log']
    headers = [str(cell.value).lower() for cell in sheet[1]]
    rows = [
        dict(zip(headers, [cell.value for cell in row]))
        for row in sheet.iter_rows(min_row=2)
    ]
    return [row for row in rows if row.get('id') is not None], manifest


def _make_backup_user(username='backup-user'):
    user = User.objects.create_user(
        username=username,
        email=f'{username}@example.com',
        password='backuppass123',
        is_staff=True,
    )
    user.user_permissions.add(
        Permission.objects.get(codename='view_reports'),
        Permission.objects.get(codename='download_backup'),
    )
    return user


@override_settings(DLUX_CONFIG=ACTIVITY_BACKUP_CONFIG)
class WriteBackupZipWindowTests(TestCase):
    def setUp(self):
        self.actor = User.objects.create_superuser(
            username='backup-admin',
            email='backup-admin@example.com',
            password='adminpass123',
        )

    def _build(self, window):
        buffer = io.BytesIO()
        manifest = write_backup_zip(self.actor, buffer, window=window)
        return buffer.getvalue(), manifest

    def test_window_all_includes_everything(self):
        recent, old = _make_logs()
        content, manifest = self._build('all')
        rows, zip_manifest = _zip_activity_rows(content)
        self.assertEqual({row['id'] for row in rows}, {recent.pk, old.pk})
        self.assertEqual(zip_manifest['window'], 'all')

    def test_window_week_filters_old_rows(self):
        recent, old = _make_logs()
        content, manifest = self._build('week')
        rows, zip_manifest = _zip_activity_rows(content)
        self.assertEqual({row['id'] for row in rows}, {recent.pk})
        self.assertEqual(zip_manifest['window'], 'week')
        model_counts = {item['model']: item['count'] for item in manifest['models']}
        self.assertEqual(model_counts.get('dlux.activitylog'), 1)

    def test_calendar_and_custom_period_bounds(self):
        now = timezone.make_aware(datetime(2026, 8, 3, 14, 30))

        quarter_start, _ = get_report_window_bounds('quarter', now=now)
        half_start, _ = get_report_window_bounds('half_year', now=now)
        year_start, _ = get_report_window_bounds('year', now=now)
        custom_start, custom_end = get_report_window_bounds(
            'custom',
            now=now,
            custom_start='2026-04-01',
            custom_end='2026-06-30',
        )

        self.assertEqual(quarter_start.date().isoformat(), '2026-07-01')
        self.assertEqual(half_start.date().isoformat(), '2026-07-01')
        self.assertEqual(year_start.date().isoformat(), '2026-01-01')
        self.assertEqual(custom_start.date().isoformat(), '2026-04-01')
        self.assertEqual(custom_end.date().isoformat(), '2026-07-01')

    def test_week_and_month_windows_are_current_period_to_date(self):
        """The week/month options cover the *current* calendar period, not the previous one.

        Regression: they were labelled "Last Week"/"Last Month" while the bounds
        have always started at the current week/month.
        """
        from dlux.translations import get_strings

        now = timezone.make_aware(datetime(2026, 8, 5, 14, 30))  # a Wednesday

        week_start, week_end = get_report_window_bounds('week', now=now)
        month_start, month_end = get_report_window_bounds('month', now=now)

        self.assertEqual(week_start.date().isoformat(), '2026-08-03')  # Monday of this week
        self.assertEqual(month_start.date().isoformat(), '2026-08-01')
        # Open-ended: both run through now rather than closing at a past boundary.
        self.assertIsNone(week_end)
        self.assertIsNone(month_end)

        for language, current in (('en', 'current'), ('ar', 'الحالي')):
            strings = get_strings(language)
            for key in ('user_report_window_week', 'user_report_window_month'):
                label = strings[key]
                with self.subTest(language=language, key=key):
                    self.assertIn(current, label.casefold() if language == 'en' else label)
                    self.assertNotIn('last', label.casefold())
                    self.assertNotIn('آخر', label)

    def test_active_period_labels_both_ends_of_open_ended_windows(self):
        """Open-ended windows run from their start through now, so say so.

        Regression: week/month/quarter/half-year/year have no upper bound, which
        left "Active period" showing a lone start date with no end.
        """
        now = timezone.make_aware(datetime(2026, 8, 5, 14, 30))

        month = build_reports_overview(self.actor, window='month', now=now)['period']
        self.assertEqual(month['start_label'], '2026-08-01')
        self.assertEqual(month['end_label'], '2026-08-05')
        self.assertEqual(month['range_label'], '2026-08-01 – 2026-08-05')

        year = build_reports_overview(self.actor, window='year', now=now)['period']
        self.assertEqual(year['range_label'], '2026-01-01 – 2026-08-05')

        custom = build_reports_overview(self.actor, window='custom', now=now, filters={
            'builder': '1', 'custom_start': '2026-04-01', 'custom_end': '2026-06-30',
        })['period']
        # Custom keeps its own inclusive end rather than being stretched to today.
        self.assertEqual(custom['range_label'], '2026-04-01 – 2026-06-30')

        # All-time has no bounds at all; templates fall back to the all-time label.
        self.assertEqual(build_reports_overview(self.actor, window='all', now=now)['period']['range_label'], '')

    def test_single_day_period_is_not_rendered_as_a_range(self):
        now = timezone.make_aware(datetime(2026, 8, 5, 9, 0))
        period = build_reports_overview(self.actor, window='custom', now=now, filters={
            'builder': '1', 'custom_start': '2026-08-05', 'custom_end': '2026-08-05',
        })['period']

        self.assertEqual(period['range_label'], '2026-08-05')
        self.assertEqual(period['start_label'], period['end_label'])

    def test_builder_selection_drives_totals_xlsx_and_zip_manifest(self):
        ActivityLog = apps.get_model('dlux', 'ActivityLog')
        ActivityLog.objects.create(action='CREATE', model_name='Project Entry')
        ActivityLog.objects.create(action='UPDATE', model_name='Project Entry')
        ActivityLog.objects.create(action='CREATE', model_name='Other Entry')
        criteria = {
            'builder': '1',
            'models': ['Project Entry'],
            'operations': ['CREATE'],
            'custom_start': timezone.localdate().isoformat(),
            'custom_end': timezone.localdate().isoformat(),
        }

        overview = build_reports_overview(self.actor, window='custom', filters=criteria)

        self.assertEqual(overview['current_total'], 1)
        self.assertEqual(overview['criteria']['models'], ['Project Entry'])
        self.assertEqual(overview['criteria']['operations'], ['CREATE'])
        workbook = build_model_entries_xlsx(self.actor, overview)
        self.assertIn(b'PK', workbook[:4])

        buffer = io.BytesIO()
        manifest = write_backup_zip(
            self.actor,
            buffer,
            window='custom',
            criteria=overview['criteria'],
        )
        with zipfile.ZipFile(io.BytesIO(buffer.getvalue())) as zf:
            names = zf.namelist()
            zip_manifest = json.loads(zf.read('manifest.json'))
        self.assertIn('entries.xlsx', names)
        # The archive ships the workbook and media only - never serialized JSON.
        self.assertEqual([name for name in names if name.startswith('data/')], [])
        self.assertEqual(zip_manifest['selection']['models'], ['Project Entry'])
        self.assertEqual(zip_manifest['selection']['operations'], ['CREATE'])
        self.assertEqual(manifest['report_artifacts'], ['entries.xlsx'])

    def test_builder_can_explicitly_exclude_every_model_and_operation(self):
        _make_logs()
        overview = build_reports_overview(
            self.actor,
            window='all',
            filters={'builder': '1', 'models': [], 'operations': []},
        )

        self.assertEqual(overview['current_total'], 0)
        self.assertEqual(overview['selected_model_count'], 0)
        self.assertEqual(overview['selected_operation_count'], 0)

    def test_builder_excludes_celery_infrastructure_models_and_activity(self):
        celery_model = SimpleNamespace(_meta=SimpleNamespace(
            abstract=False,
            managed=True,
            app_label='django_celery_results',
        ))
        self.assertFalse(is_report_eligible_model(celery_model))
        for identity in (
            'django_celery_results.taskresult',
            'django_celery_beat.periodictask',
            'djcelery.taskmeta',
            'Task Result',
            'Periodic Task',
        ):
            self.assertFalse(is_report_eligible_activity_model_name(identity))

        ActivityLog = apps.get_model('dlux', 'ActivityLog')
        ActivityLog.objects.create(
            action='UPDATE',
            model_key='django_celery_results.taskresult',
            model_name='Task Result',
        )
        ActivityLog.objects.create(
            action='UPDATE',
            model_key='django_celery_beat.periodictask',
            model_name='Periodic Task',
        )
        ActivityLog.objects.create(action='CREATE', model_name='Project Entry')

        overview = build_reports_overview(self.actor, window='all')
        available_keys = {item['key'] for item in overview['available_models']}

        self.assertEqual(overview['all_total'], 1)
        self.assertNotIn('django_celery_results.taskresult', available_keys)
        self.assertNotIn('django_celery_beat.periodictask', available_keys)

    def test_builder_supports_model_and_config_level_developer_exclusions(self):
        meta = SimpleNamespace(
            abstract=False,
            managed=True,
            auto_created=False,
            proxy=False,
            swapped=False,
            app_label='documents',
            label_lower='documents.internalrecord',
            model_name='internalrecord',
            object_name='InternalRecord',
            verbose_name='Internal record',
            verbose_name_plural='Internal records',
        )
        model = SimpleNamespace(_meta=meta, dlux_report=False)
        self.assertFalse(is_report_eligible_model(model))
        with override_settings(DLUX_CONFIG={
            'reports': {
                'include_models': ['documents.internalrecord'],
                'include_activity': ['documents.internalrecord'],
            },
        }), patch('dlux.reports.eligibility.resolve_model_by_name', return_value=model):
            self.assertFalse(is_report_eligible_model(model))
            self.assertFalse(
                is_report_eligible_activity_model_name('documents.internalrecord')
            )

        del model.dlux_report
        with override_settings(DLUX_CONFIG={
            'reports': {
                'exclude_models': ['documents.internalrecord'],
                'include_activity': ['documents.internalrecord'],
            },
        }), patch('dlux.reports.eligibility.resolve_model_by_name', return_value=model):
            self.assertFalse(is_report_eligible_model(model))
            self.assertFalse(
                is_report_eligible_activity_model_name('documents.internalrecord')
            )

        ActivityLog = apps.get_model('dlux', 'ActivityLog')
        ActivityLog.objects.create(
            action='CREATE',
            model_key='documents.internalrecord',
            model_name='Internal record',
        )
        ActivityLog.objects.create(action='CREATE', model_name='Project Entry')
        with override_settings(DLUX_CONFIG={
            'reports': {
                'exclude_models': ['documents.internalrecord'],
                'include_activity': ['documents.internalrecord'],
            },
        }), patch(
            'dlux.reports.eligibility.resolve_model_by_name',
            side_effect=lambda value: model if value == 'documents.internalrecord' else None,
        ):
            overview = build_reports_overview(self.actor, window='all')

        self.assertEqual(overview['all_total'], 1)
        self.assertNotIn(
            'documents.internalrecord',
            {item['key'] for item in overview['available_models']},
        )

    def test_report_action_row_uses_full_logical_width_for_rtl_alignment(self):
        stylesheet = (
            Path(__file__).resolve().parents[1]
            / 'static' / 'dlux' / 'reports' / 'css' / 'overview.css'
        ).read_text(encoding='utf-8')
        action_rule = stylesheet.split('.dlux-report-builder-submit {', 1)[1].split('}', 1)[0]

        self.assertIn('display: flex;', action_rule)
        self.assertIn('width: 100%;', action_rule)
        self.assertIn('justify-content: flex-end;', action_rule)

    def test_report_templates_do_not_use_django_comments(self):
        template_root = Path(__file__).resolve().parents[1] / 'templates' / 'dlux'
        paths = (
            template_root / 'reports' / 'overview.html',
            template_root / 'reports' / 'print.html',
            template_root / 'tables' / 'pair_table.html',
            template_root / 'users' / '_user_report_activity.html',
            template_root / 'users' / '_user_report_window.html',
            template_root / 'users' / 'user_report_modal.html',
        )

        for path in paths:
            with self.subTest(path=path):
                source = path.read_text(encoding='utf-8')
                self.assertNotIn('{#', source)
                self.assertNotIn('#}', source)

    def test_period_selection_reloads_the_builder_from_external_js(self):
        """Changing the period must re-query, since every figure/export depends on it."""
        script = (
            Path(__file__).resolve().parents[1]
            / 'static' / 'dlux' / 'reports' / 'js' / 'overview.js'
        ).read_text(encoding='utf-8')
        handler = script.split("windowSelect.addEventListener('change'", 1)[1].split('});', 1)[0]

        self.assertIn('submitBuilder()', handler)
        # A half-filled custom range must not auto-submit.
        self.assertIn("windowSelect.value !== 'custom' || bothCustomDatesFilled()", handler)
        self.assertIn('requestSubmit', script)

    def test_print_assets_preserve_the_a4_grid_and_remeasure_charts(self):
        static_root = Path(__file__).resolve().parents[1] / 'static' / 'dlux' / 'reports'
        stylesheet = (static_root / 'css' / 'print.css').read_text(encoding='utf-8')
        script = (static_root / 'js' / 'print.js').read_text(encoding='utf-8')
        print_rules = stylesheet.split('@media print {', 1)[1]

        self.assertIn('grid-template-columns: repeat(4, minmax(0, 1fr));', print_rules)
        self.assertIn('grid-template-columns: repeat(2, minmax(0, 1fr));', print_rules)
        self.assertIn('.dlux-report-doc-page-two {', print_rules)
        self.assertIn('break-before: page;', print_rules)
        self.assertIn('width: 100% !important;', print_rules)
        self.assertIn('chart.resize(width, height);', script)
        self.assertIn('reverse: false,', script)

    def test_backup_status_is_narrated_by_exactly_one_element(self):
        """The bar shows progress; one status line carries the words.

        Regression: completion wrote the same message into both the progress
        label and the note, so every terminal message rendered twice.
        """
        root = Path(__file__).resolve().parents[1]
        script = (root / 'static' / 'dlux' / 'reports' / 'js' / 'overview.js').read_text(encoding='utf-8')
        template = (root / 'templates' / 'dlux' / 'reports' / 'overview.html').read_text(encoding='utf-8')

        self.assertNotIn('data-reports-backup-progress-label', template)
        self.assertNotIn('data-reports-backup-progress-label', script)
        # setProgress moves the bar only - it takes no message argument.
        self.assertIn('function setProgress(value, terminalTone)', script)
        completed = script.split("if (data.status === 'completed'", 1)[1].split('} else if', 1)[0]
        self.assertEqual(completed.count('setNote'), 0)
        self.assertEqual(completed.count('finish('), 1)

    def test_backup_button_label_describes_creating_not_downloading(self):
        template = (
            Path(__file__).resolve().parents[1]
            / 'templates' / 'dlux' / 'reports' / '_backup_action.html'
        ).read_text(encoding='utf-8')
        action = template.split('id="reports-backup-btn"', 1)[1].split('</button>', 1)[0]

        self.assertIn('DLUX_STRINGS.reports_backup_create', action)
        self.assertNotIn('DLUX_STRINGS.reports_backup_zip', action)
        from dlux.translations import get_strings

        for language in ('en', 'ar'):
            strings = get_strings(language)
            self.assertTrue(strings['reports_backup_create'])
            self.assertTrue(strings['reports_backup_download_latest'])
            self.assertNotEqual(
                strings['reports_backup_create'],
                strings['reports_backup_download_latest'],
            )

    def test_builder_excludes_dlux_system_audit_and_legacy_backup_activity(self):
        self.assertFalse(is_report_eligible_activity_model_name('Dlux System Backup'))
        self.assertFalse(is_report_eligible_activity_model_name('Dlux System Restore'))

        ActivityLog = apps.get_model('dlux', 'ActivityLog')
        ActivityLog.objects.create(
            created_by=self.actor,
            action='RESTORE',
            category='system',
            model_name='Dlux System Backup',
        )
        ActivityLog.objects.create(
            created_by=self.actor,
            action='UPDATE',
            category='system',
            model_name='Project Entry',
        )
        ActivityLog.objects.create(
            created_by=self.actor,
            action='PASSWORD_RESET',
            category='audit',
            model_name='Project Entry',
        )
        ActivityLog.objects.create(
            created_by=self.actor,
            action='RESTORE',
            category='user',
            model_name='Dlux System Backup',
        )
        project_log = ActivityLog.objects.create(
            created_by=self.actor,
            action='CREATE',
            category='user',
            model_name='Project Entry',
        )

        overview = build_reports_overview(self.actor, window='all')
        available_keys = {item['key'] for item in overview['available_models']}
        available_operations = {item['key'] for item in overview['available_actions']}

        self.assertEqual(overview['all_total'], 1)
        self.assertEqual(list(overview['activity_qs']), [project_log])
        self.assertNotIn('Dlux System Backup', available_keys)
        self.assertNotIn('RESTORE', available_operations)
        self.assertNotIn('PASSWORD_RESET', available_operations)

        from dlux.reports.users import build_user_report
        user_report = build_user_report(self.actor, actor=self.actor, window='all')
        self.assertEqual(user_report['summary']['activity_count'], 1)

    def test_invalid_window_defaults_to_all(self):
        recent, old = _make_logs()
        content, manifest = self._build('bogus')
        rows, _ = _zip_activity_rows(content)
        self.assertEqual(manifest['window'], 'all')
        self.assertEqual(len(rows), 2)

    def test_record_folder_prefers_business_number_and_keeps_pk(self):
        record = NumberedBackupRecord(pk=37, number='2000 / A')

        self.assertEqual(
            backup_record_folder(record),
            'number-2000-A',
        )

    @override_settings(DLUX_CONFIG={
        'reports': {
            'backup_label_fields': {
                'backup_tests.configuredbackuprecord': 'case_reference',
            },
        },
    })
    def test_record_folder_supports_explicit_per_model_label_field(self):
        record = ConfiguredBackupRecord(pk=8, case_reference='../CASE\\42')

        self.assertEqual(
            backup_record_folder(record),
            'case-reference-CASE-42',
        )


@override_settings(DLUX_CONFIG=ACTIVITY_BACKUP_CONFIG)
class ModelEntriesExportTests(TestCase):
    """The XLSX export carries real model rows, not the report's aggregate figures."""

    def setUp(self):
        self.actor = User.objects.create_superuser(
            username='entries-admin',
            email='entries-admin@example.com',
            password='adminpass123',
        )

    def test_export_sheets_hold_actual_rows_not_aggregate_report_data(self):
        ActivityLog = apps.get_model('dlux', 'ActivityLog')
        first = ActivityLog.objects.create(
            created_by=self.actor, action='CREATE', model_name='Project Entry', number='INV-1',
        )
        second = ActivityLog.objects.create(
            created_by=self.actor, action='UPDATE', model_name='Project Entry', number='INV-2',
        )

        overview = build_reports_overview(self.actor, window='all')
        workbook = _load_workbook(build_model_entries_xlsx(self.actor, overview))

        sheet = workbook[workbook.sheetnames[1]]
        headers = [str(cell.value).lower() for cell in sheet[1]]
        self.assertIn('document number', headers)
        numbers = {row[headers.index('document number')].value for row in sheet.iter_rows(min_row=2)}
        self.assertEqual(numbers, {first.number, second.number})
        # The aggregate sheets the old export produced must be gone.
        self.assertNotIn('By User', workbook.sheetnames)
        self.assertNotIn('By Day', workbook.sheetnames)
        self.assertEqual(sheet.max_row, 3)

    def test_export_honors_period_and_operation_selection(self):
        ActivityLog = apps.get_model('dlux', 'ActivityLog')
        row = dict(created_by=self.actor, model_name='Activity Log', model_key='dlux.activitylog')
        kept = ActivityLog.objects.create(action='CREATE', number='KEEP', **row)
        dropped = ActivityLog.objects.create(action='UPDATE', number='DROP', **row)
        stale = ActivityLog.objects.create(action='CREATE', number='OLD', **row)
        ActivityLog.objects.filter(pk=stale.pk).update(
            created_at=timezone.now() - timedelta(days=60),
        )

        overview = build_reports_overview(self.actor, window='week', filters={
            'builder': '1',
            'models': ['dlux.activitylog'],
            'operations': ['CREATE'],
        })
        workbook = _load_workbook(build_model_entries_xlsx(self.actor, overview))
        sheet = workbook[workbook.sheetnames[1]]
        headers = [str(cell.value).lower() for cell in sheet[1]]
        numbers = {row[headers.index('document number')].value for row in sheet.iter_rows(min_row=2)}

        self.assertEqual(numbers, {kept.number})
        self.assertNotIn(dropped.number, numbers)
        self.assertNotIn(stale.number, numbers)

    def test_export_without_selected_models_still_produces_info_sheet(self):
        _make_logs()
        overview = build_reports_overview(self.actor, window='all', filters={
            'builder': '1', 'models': [], 'operations': [],
        })
        workbook = _load_workbook(build_model_entries_xlsx(self.actor, overview))

        self.assertEqual(len(workbook.sheetnames), 1)
        values = {str(row[0]) for row in workbook.active.iter_rows(min_row=1, values_only=True)}
        self.assertIn('No models were selected, so no entry sheets were produced.', values)

    def test_export_field_list_drops_credential_bearing_columns(self):
        class SecretiveRecord(models.Model):
            name = models.CharField(max_length=50)
            password = models.CharField(max_length=128)
            api_key_value = models.CharField(max_length=128)
            reset_token = models.CharField(max_length=128)

            class Meta:
                app_label = 'backup_tests'
                managed = False

        exported = {field.name for field in get_report_entry_fields(SecretiveRecord)}

        self.assertIn('name', exported)
        self.assertNotIn('password', exported)
        self.assertNotIn('api_key_value', exported)
        self.assertNotIn('reset_token', exported)

    @override_settings(DLUX_CONFIG={
        'reports': {
            'include_models': ['dlux.activitylog'],
            'entries_row_limit': 1,
        },
    })
    def test_export_row_limit_is_enforced_and_reported(self):
        ActivityLog = apps.get_model('dlux', 'ActivityLog')
        for index in range(3):
            ActivityLog.objects.create(
                created_by=self.actor, action='CREATE', model_name='Project Entry', number=f'N{index}',
            )

        overview = build_reports_overview(self.actor, window='all')
        workbook = _load_workbook(build_model_entries_xlsx(self.actor, overview))

        self.assertEqual(workbook[workbook.sheetnames[1]].max_row, 2)
        info_values = {
            str(cell)
            for row in workbook.active.iter_rows(values_only=True)
            for cell in row
        }
        self.assertIn('Yes', info_values)

    def test_report_zip_packs_workbook_and_media_without_json(self):
        """The periodic ZIP is a deliverable: workbook + media, no serialized rows."""
        ActivityLog = apps.get_model('dlux', 'ActivityLog')
        ActivityLog.objects.create(
            created_by=self.actor, action='CREATE', model_name='Invoice', number='INV-9',
        )

        buffer = io.BytesIO()
        manifest = write_backup_zip(self.actor, buffer, window='all')
        with zipfile.ZipFile(io.BytesIO(buffer.getvalue())) as zf:
            names = zf.namelist()
            workbook = _load_workbook(zf.read('entries.xlsx'))

        self.assertEqual(sorted(names), ['entries.xlsx', 'manifest.json'])
        self.assertEqual([name for name in names if name.endswith('.json') and name != 'manifest.json'], [])
        self.assertEqual(manifest['report_artifacts'], ['entries.xlsx'])
        self.assertIn('Activity Log', workbook.sheetnames)
        # ActivityLog carries no uploads, so there is nothing to place under files/.
        self.assertEqual(manifest['files'], [])
        self.assertEqual(manifest['missing_files'], [])

    @override_settings(DLUX_CONFIG={'reports': {'include_models': ['dlux.profile']}})
    def test_report_zip_packs_referenced_media_alongside_the_workbook(self):
        from django.core.files.uploadedfile import SimpleUploadedFile

        profile = self.actor.profile
        profile.profile_picture.save(
            'avatar.png', SimpleUploadedFile('avatar.png', b'binary-image-bytes'), save=True,
        )
        self.addCleanup(lambda: default_storage.delete(profile.profile_picture.name))

        buffer = io.BytesIO()
        manifest = write_backup_zip(self.actor, buffer, window='all')
        with zipfile.ZipFile(io.BytesIO(buffer.getvalue())) as zf:
            names = zf.namelist()
            packed = [name for name in names if name.startswith('files/')]
            payload = zf.read(packed[0])

        self.assertIn('entries.xlsx', names)
        self.assertEqual(len(packed), 1)
        self.assertTrue(packed[0].startswith('files/dlux/profile/'))
        self.assertTrue(packed[0].endswith('/profile_picture/avatar.png'))
        self.assertEqual(payload, b'binary-image-bytes')
        self.assertEqual(manifest['files'][0]['field'], 'profile_picture')
        self.assertEqual(manifest['files'][0]['name'], profile.profile_picture.name)
        self.assertEqual(manifest['missing_files'], [])

    def test_chart_payload_orders_days_and_caps_categorical_slots(self):
        ActivityLog = apps.get_model('dlux', 'ActivityLog')
        for index in range(9):
            ActivityLog.objects.create(
                created_by=self.actor,
                action=f'OP{index}',
                model_name='Project Entry',
            )

        overview = build_reports_overview(self.actor, window='all')
        payload = build_report_chart_data(overview)

        self.assertLessEqual(len(payload['actions']), 8)
        self.assertEqual(payload['actions'][-1]['label'], 'Other')
        day_labels = [item['label'] for item in payload['days']]
        self.assertEqual(day_labels, sorted(day_labels))


@override_settings(DLUX_CONFIG=ACTIVITY_BACKUP_CONFIG)
class BackupViewTests(TestCase):
    def setUp(self):
        settings_obj = apps.get_model('dlux', 'SystemSettings').load()
        settings_obj.is_configured = True
        settings_obj.save(update_fields=['is_configured'])
        self.user = _make_backup_user()
        self.client = Client()
        self.client.login(username='backup-user', password='backuppass123')

    def test_sync_backup_zip_honors_window(self):
        recent, old = _make_logs()
        response = self.client.get(reverse('reports_backup_zip'), {'window': 'week'})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/zip')
        content = b''.join(response.streaming_content)
        rows, manifest = _zip_activity_rows(content)
        # The EXPORT row logged for this download postdates the snapshot, so only
        # `recent` is guaranteed; `old` must be filtered out by the week window.
        self.assertIn(recent.pk, {row['id'] for row in rows})
        self.assertNotIn(old.pk, {row['id'] for row in rows})
        self.assertEqual(manifest['window'], 'week')

    def test_start_falls_back_to_sync_without_worker(self):
        _make_logs()
        response = self.client.post(reverse('reports_backup_start'), {
            'window': 'month',
            'builder': '1',
            'models': ['Project Entry'],
            'operations': ['CREATE'],
        })
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['ok'])
        self.assertFalse(payload['async'])
        self.assertIn('window=month', payload['download_url'])
        self.assertIn('models=Project+Entry', payload['download_url'])
        self.assertIn('operations=CREATE', payload['download_url'])
        # No orphaned pending row is left behind on fallback.
        ReportBackup = apps.get_model('dlux', 'ReportBackup')
        self.assertEqual(ReportBackup.objects.count(), 0)

    def test_xlsx_export_downloads_model_entries(self):
        ActivityLog = apps.get_model('dlux', 'ActivityLog')
        entry = ActivityLog.objects.create(
            created_by=self.user, action='CREATE', model_name='Project Entry', number='ROW-1',
        )

        response = self.client.get(reverse('reports_overview_xlsx'), {'window': 'all'})

        self.assertEqual(response.status_code, 200)
        self.assertIn('dlux-entries-', response['Content-Disposition'])
        workbook = _load_workbook(response.content)
        sheet = workbook[workbook.sheetnames[1]]
        headers = [str(cell.value).lower() for cell in sheet[1]]
        numbers = {row[headers.index('document number')].value for row in sheet.iter_rows(min_row=2)}
        self.assertIn(entry.number, numbers)

    def test_print_report_renders_charts_and_selection(self):
        _make_logs()

        response = self.client.get(reverse('reports_print'), {'window': 'all'})

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'dlux/reports/print.html')
        self.assertContains(response, 'dlux-report-chart-data')
        self.assertContains(response, 'data-dlux-report-chart="days"')
        self.assertContains(response, 'data-dlux-report-chart="actions"')
        self.assertContains(response, 'dlux-report-doc-page-two')
        self.assertIn('days', response.context['chart_data'])
        self.assertIn(reverse('reports_overview'), response.context['builder_url'])

    def test_report_pages_carry_no_inline_handlers_or_unnonced_scripts(self):
        """CSP: behaviour lives in external, nonce'd scripts on both report pages."""
        import re

        _make_logs()
        for url in (reverse('reports_overview'), reverse('reports_print')):
            html = self.client.get(url).content.decode()
            with self.subTest(url=url):
                self.assertEqual(re.findall(r'\son[a-z]+\s*=\s*["\']', html), [])
                for tag in re.findall(r'<script\b[^>]*>', html):
                    # json_script data blocks are not executable, so they need no nonce.
                    if 'type="application/json"' in tag:
                        continue
                    self.assertIn('nonce=', tag)
                    self.assertIn('src=', tag)

    def test_print_report_requires_reports_permission(self):
        plain = User.objects.create_user(
            username='no-reports', password='plainpass123', is_staff=True,
        )
        client = Client()
        client.login(username='no-reports', password='plainpass123')

        self.assertEqual(client.get(reverse('reports_print')).status_code, 403)
        self.assertEqual(client.get(reverse('reports_overview_xlsx')).status_code, 403)

    def test_start_requires_backup_permission(self):
        plain = User.objects.create_user(
            username='plain-user', password='plainpass123', is_staff=True,
        )
        client = Client()
        client.login(username='plain-user', password='plainpass123')
        response = client.post(reverse('reports_backup_start'), {'window': 'all'})
        self.assertEqual(response.status_code, 403)

    def test_async_start_persists_normalized_builder_criteria(self):
        _make_logs()
        with patch('dlux.views.reports.dispatch_report_backup', return_value=True):
            response = self.client.post(reverse('reports_backup_start'), {
                'window': 'custom',
                'custom_start': timezone.localdate().isoformat(),
                'custom_end': timezone.localdate().isoformat(),
                'builder': '1',
                'models': ['Project Entry', 'not.allowed'],
                'operations': ['CREATE', 'not.allowed'],
            })

        self.assertTrue(response.json()['async'])
        backup = apps.get_model('dlux', 'ReportBackup').objects.get()
        self.assertEqual(backup.criteria['models'], ['Project Entry'])
        self.assertEqual(backup.criteria['operations'], ['CREATE'])
        self.assertEqual(backup.criteria['custom_start'], timezone.localdate().isoformat())

    def test_start_reuses_active_backup_instead_of_queueing_duplicate(self):
        ReportBackup = apps.get_model('dlux', 'ReportBackup')
        active = ReportBackup.objects.create(user=self.user, window='all')

        with patch('dlux.views.reports.dispatch_report_backup') as dispatch:
            response = self.client.post(reverse('reports_backup_start'), {'window': 'week'})

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['async'])
        self.assertTrue(payload['reused'])
        self.assertEqual(payload['token'], active.token)
        self.assertEqual(ReportBackup.objects.count(), 1)
        dispatch.assert_not_called()

    def test_overview_resumes_active_backup_and_links_latest_completed(self):
        ReportBackup = apps.get_model('dlux', 'ReportBackup')
        completed = ReportBackup.objects.create(
            user=self.user,
            window='week',
            status=ReportBackup.STATUS_COMPLETED,
            file_path='dlux_backups/completed.zip',
            completed_at=timezone.now(),
        )
        active = ReportBackup.objects.create(user=self.user, window='all')

        response = self.client.get(reverse('reports_overview'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            reverse('reports_backup_status', args=[active.token]),
        )
        self.assertContains(
            response,
            reverse('reports_backup_download', args=[completed.token]),
        )
        self.assertContains(response, 'data-report-custom-period')
        self.assertContains(response, 'data-report-choice-panel="models"')
        self.assertContains(response, 'dlux-table-shell')

    def test_run_status_and_download_flow(self):
        recent, old = _make_logs()
        ReportBackup = apps.get_model('dlux', 'ReportBackup')
        backup = ReportBackup.objects.create(user=self.user, window='week')
        with tempfile.TemporaryDirectory() as media_root:
            with override_settings(MEDIA_ROOT=media_root):
                run_report_backup(backup.pk)
                backup.refresh_from_db()
                self.assertEqual(backup.status, ReportBackup.STATUS_COMPLETED)
                self.assertTrue(backup.file_path)
                self.assertTrue(default_storage.exists(backup.file_path))
                self.assertGreater(backup.file_size, 0)
                self.assertEqual(backup.progress_percent, 100)
                self.assertTrue(backup.progress_message)

                DluxNotification = apps.get_model('dlux', 'DluxNotification')
                self.assertTrue(DluxNotification.objects.filter(
                    source_model_key='dlux.reportbackup',
                    source_object_id=str(backup.pk),
                    action='backup_progress',
                    metadata__locked=False,
                ).exists())
                self.assertTrue(DluxNotification.objects.filter(
                    source_model_key='dlux.reportbackup',
                    source_object_id=str(backup.pk),
                    action='backup_completed',
                ).exists())

                status = self.client.get(
                    reverse('reports_backup_status', args=[backup.token])
                )
                self.assertEqual(status.status_code, 200)
                payload = status.json()
                self.assertEqual(payload['status'], 'completed')
                self.assertEqual(payload['progress_percent'], 100)
                self.assertIn('download_url', payload)
                self.assertIn('no-cache', status['Cache-Control'])

                download = self.client.get(payload['download_url'])
                self.assertEqual(download.status_code, 200)
                self.assertEqual(download['Content-Type'], 'application/zip')
                content = b''.join(download.streaming_content)
                rows, manifest = _zip_activity_rows(content)
                self.assertIn(recent.pk, {row['id'] for row in rows})
                self.assertNotIn(old.pk, {row['id'] for row in rows})

                # Another permitted user cannot see someone else's backup.
                other = _make_backup_user('backup-other')
                other_client = Client()
                other_client.login(username='backup-other', password='backuppass123')
                denied = other_client.get(
                    reverse('reports_backup_status', args=[backup.token])
                )
                self.assertEqual(denied.status_code, 404)
