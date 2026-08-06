import io
import json
import logging
import re
import shutil
import tempfile
import unicodedata
import zipfile
from collections import Counter, defaultdict
from datetime import date, datetime, time, timedelta
from decimal import Decimal
from io import BytesIO

from django.apps import apps
from django.conf import settings
from django.core.cache import cache
from django.core.serializers.json import Serializer as JsonSerializer
from django.core.files import File
from django.core.files.storage import default_storage
from django.db import models
from django.db.models import Count, Q
from django.db.models.functions import TruncDate
from django.utils import timezone

from .translations import get_current_language_code, get_strings
from .utils import (
    get_user_scope,
    is_central_staff,
    is_global_staff,
    is_scope_enabled,
    log_user_action,
    normalize_activity_log_model_key,
    resolve_model_by_name,
    translate_activity_log_model_name,
)


logger = logging.getLogger("dlux")

REPORT_WINDOWS = {"week", "month", "quarter", "half_year", "year", "custom", "all"}
REPORT_ACTIVITY_CATEGORY = "user"
REPORT_OVERVIEW_CACHE_SCHEMA_VERSION = 2
REPORT_ENTRIES_ROW_LIMIT = 20000
REPORT_CHART_TOP_N = 12
# The operation mix is the one categorical chart: it may never exceed the eight
# validated hue slots, so the tail folds into a single "Other" slice.
REPORT_CHART_CATEGORICAL_TOP_N = 7
_DLUX_REPORT_EXCLUDED_KEYS = {
    "auth",
    "authentication",
    "session",
    "sessions",
    "systemsettings",
    "system_settings",
    "scope",
    "scopesettings",
    "scope_settings",
    "profile",
    "user_profile",
    "trusteddevice",
    "trusted_device",
    "userknowndevice",
    "known_device",
    "known_devices",
    "userpresencesession",
    "presence_session",
    "presence_sessions",
    "publicregistration",
    "public_registration",
    "preferences",
    "password",
    "totp",
    "otp",
    "backup_code",
    "backup_codes",
    # Dlux-internal meta actions (not project work; classified as system interactions)
    "dlux_system_backup",
    "dlux_system_restore",
    "dlux_reports_backup",
    "reports_backup",
}

# Verbose names of the high-frequency operational tracking models that are no longer
# logged (see signals.EXCLUDED_MODELS). Historical rows may still exist; use
# exclude_log_noise() to keep them out of activity feeds and reports entirely.
LOG_NOISE_MODEL_NAMES = ("Trusted Device", "Known Device", "Presence Session")


def exclude_log_noise(queryset):
    """Drop operational tracking-model rows (presence/device churn) from a log queryset."""
    return queryset.exclude(model_name__in=LOG_NOISE_MODEL_NAMES)
_EXCLUDED_APP_LABELS = {
    "admin",
    "auth",
    "contenttypes",
    "sessions",
    "messages",
    "staticfiles",
}
_CELERY_REPORT_APP_LABELS = {
    "celery",
    "djcelery",
    "django_celery_beat",
    "django_celery_results",
}
_CELERY_REPORT_MODEL_KEYS = {
    "task_result",
    "taskresult",
    "group_result",
    "groupresult",
    "chord_counter",
    "chordcounter",
    "periodic_task",
    "periodictask",
    "periodic_tasks",
    "periodictasks",
    "crontab_schedule",
    "crontabschedule",
    "interval_schedule",
    "intervalschedule",
    "solar_schedule",
    "solarschedule",
    "clocked_schedule",
    "clockedschedule",
    "task_meta",
    "taskmeta",
    "task_set_meta",
    "tasksetmeta",
}


def normalize_report_window(value):
    value = str(value or "week").strip().lower()
    return value if value in REPORT_WINDOWS else "week"


def normalize_backup_window(value):
    """Like normalize_report_window but defaults to 'all' (the historical backup scope)."""
    value = str(value or "all").strip().lower()
    return value if value in REPORT_WINDOWS else "all"


def _parse_report_date(value):
    if isinstance(value, date):
        return value
    try:
        return date.fromisoformat(str(value or "").strip())
    except (TypeError, ValueError):
        return None


def _reports_config():
    config = getattr(settings, "DLUX_CONFIG", {}).get("reports", {})
    return config if isinstance(config, dict) else {}


def get_reports_overview_cache_seconds():
    """Optional read-through cache TTL for the expensive report overview stats.

    Defaults to 0 to keep the overview exact out of the box. Deployments with a
    shared Django cache/Redis can set DLUX_CONFIG['reports']['overview_cache_seconds']
    to a small value to absorb repeated dashboard loads.
    """
    try:
        value = int(_reports_config().get("overview_cache_seconds", 0))
    except (TypeError, ValueError):
        value = 0
    return max(0, min(value, 3600))


def _normalized_config_set(key):
    values = _reports_config().get(key, [])
    if isinstance(values, str):
        values = [values]
    return {normalize_activity_log_model_key(value) for value in values or [] if str(value or "").strip()}


def _model_keys(model):
    meta = model._meta
    values = {
        meta.label_lower,
        f"{meta.app_label}.{meta.model_name}",
        meta.model_name,
        meta.object_name,
        str(meta.verbose_name),
        str(meta.verbose_name_plural),
    }
    return {normalize_activity_log_model_key(value) for value in values if str(value or "").strip()}


def _model_is_section(model):
    marker = getattr(model, "is_section", None)
    if isinstance(marker, bool):
        return marker
    if marker is not None:
        return True
    return bool(getattr(model._meta, "is_section", False))


def _is_celery_report_identity(value):
    normalized = normalize_activity_log_model_key(value)
    return (
        normalized in _CELERY_REPORT_MODEL_KEYS
        or normalized.startswith("django_celery_beat_")
        or normalized.startswith("django_celery_results_")
        or normalized.startswith("djcelery_")
        or normalized.startswith("celery_")
    )


def _is_celery_report_model(model):
    app_label = str(model._meta.app_label or "").lower()
    return app_label in _CELERY_REPORT_APP_LABELS or app_label.startswith("django_celery_")


def is_report_eligible_model(model):
    if model is None:
        return False
    meta = model._meta
    if (
        meta.abstract
        or not meta.managed
        or getattr(meta, "auto_created", False)
        or getattr(meta, "proxy", False)
        or getattr(meta, "swapped", False)
    ):
        return False
    if _is_celery_report_model(model):
        return False
    if getattr(model, "dlux_report", None) is False:
        return False
    keys = _model_keys(model)
    if keys & _normalized_config_set("exclude_models"):
        return False
    if keys & _normalized_config_set("include_models"):
        return True
    if meta.app_label in _EXCLUDED_APP_LABELS:
        return False
    if meta.app_label == "dlux":
        return _model_is_section(model)
    return True


def is_report_eligible_activity_model_name(model_name):
    raw = str(model_name or "").strip()
    if not raw:
        return False
    # Non-ASCII verbose names (e.g. translated/Arabic model labels stored by the
    # activity logger) normalize to an empty key. Don't reject them outright — the
    # key-based checks below simply won't match, and eligibility falls through to
    # model resolution so the decision is made on the underlying model instead of
    # on whether its label happens to be ASCII.
    normalized = normalize_activity_log_model_key(raw)
    if _is_celery_report_identity(raw):
        return False
    if normalized and normalized in _DLUX_REPORT_EXCLUDED_KEYS:
        return False
    model = resolve_model_by_name(raw)
    if model is not None and not is_report_eligible_model(model):
        return False
    if normalized in _normalized_config_set("exclude_activity"):
        return False
    if normalized in _normalized_config_set("include_activity"):
        return True
    if model is not None:
        return True
    return True


def get_report_eligible_models():
    return [model for model in apps.get_models() if is_report_eligible_model(model)]


def _week_start_index():
    value = _reports_config().get("week_start", 0)
    try:
        value = int(value)
    except (TypeError, ValueError):
        value = 0
    return value if 0 <= value <= 6 else 0


def get_report_window_bounds(window, *, now=None, custom_start=None, custom_end=None):
    window = normalize_report_window(window)
    if window == "all":
        return None, None
    now = timezone.localtime(now or timezone.now())
    timezone_value = timezone.get_current_timezone()
    if window == "custom":
        start_date = _parse_report_date(custom_start) or now.date()
        end_date = _parse_report_date(custom_end) or start_date
        if end_date < start_date:
            start_date, end_date = end_date, start_date
        start = timezone.make_aware(datetime.combine(start_date, time.min), timezone_value)
        end = timezone.make_aware(datetime.combine(end_date + timedelta(days=1), time.min), timezone_value)
        return start, end
    if window == "year":
        return now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0), None
    if window == "half_year":
        month = 1 if now.month <= 6 else 7
        return now.replace(month=month, day=1, hour=0, minute=0, second=0, microsecond=0), None
    if window == "quarter":
        month = ((now.month - 1) // 3) * 3 + 1
        return now.replace(month=month, day=1, hour=0, minute=0, second=0, microsecond=0), None
    if window == "month":
        start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        return start, None
    week_start = _week_start_index()
    start_date = now.date() - timedelta(days=(now.weekday() - week_start) % 7)
    start = timezone.make_aware(datetime.combine(start_date, time.min), timezone.get_current_timezone())
    return start, None


def get_previous_period_bounds(window, *, now=None, custom_start=None, custom_end=None):
    now = timezone.localtime(now or timezone.now())
    current_start, current_end = get_report_window_bounds(
        window,
        now=now,
        custom_start=custom_start,
        custom_end=custom_end,
    )
    if current_start is None:
        return None, None
    comparison_end = current_start
    duration = (current_end or now) - current_start
    return comparison_end - duration, comparison_end


def get_previous_week_bounds(*, now=None):
    current_start, _ = get_report_window_bounds("week", now=now)
    previous_start = current_start - timedelta(days=7)
    return previous_start, current_start


def _base_activity_queryset():
    ActivityLog = apps.get_model("dlux", "ActivityLog")
    manager = getattr(ActivityLog, "all_objects", ActivityLog._default_manager)
    qs = manager.filter(deleted_at__isnull=True) if hasattr(ActivityLog, "deleted_at") else manager.all()
    return qs.select_related("created_by", "created_by__profile__scope", "scope")


def apply_report_scope(queryset, actor):
    if not is_scope_enabled():
        return queryset
    if getattr(actor, "is_superuser", False) or is_global_staff(actor):
        return queryset
    if is_central_staff(actor):
        return queryset.filter(scope__isnull=True)
    actor_scope = get_user_scope(actor)
    if actor_scope is not None:
        return queryset.filter(scope=actor_scope)
    return queryset.none()


def apply_report_window(queryset, window, *, now=None, custom_start=None, custom_end=None):
    start, end = get_report_window_bounds(
        window,
        now=now,
        custom_start=custom_start,
        custom_end=custom_end,
    )
    if start is not None:
        queryset = queryset.filter(created_at__gte=start)
    if end is not None:
        queryset = queryset.filter(created_at__lt=end)
    return queryset


def get_visible_report_activity(actor, *, window="all", now=None, custom_start=None, custom_end=None):
    qs = apply_report_scope(_base_activity_queryset(), actor)
    qs = apply_report_window(
        qs,
        window,
        now=now,
        custom_start=custom_start,
        custom_end=custom_end,
    )
    return qs.order_by("-created_at")


def activity_report_key(model_key, model_name):
    """Locale-independent identity for an activity row.

    Prefer the stable ``app_label.model_name`` key; fall back to the (possibly
    translated) display label for legacy rows and non-model events. Grouping and
    eligibility must run on this, never on the raw ``model_name`` label, which
    changes with the active language.
    """
    key = str(model_key or "").strip()
    if key:
        return key
    return str(model_name or "").strip()


def log_report_key(log):
    """``activity_report_key`` for a UserActivityLog instance."""
    return activity_report_key(getattr(log, "model_key", None), getattr(log, "model_name", None))


def filter_report_eligible_activity(queryset):
    queryset = queryset.filter(category=REPORT_ACTIVITY_CATEGORY)
    eligible_keys = set()
    eligible_legacy_names = set()
    identity_rows = queryset.select_related(None).order_by().values_list("model_key", "model_name").distinct()
    for model_key, model_name in identity_rows:
        report_key = activity_report_key(model_key, model_name)
        if not report_key or not is_report_eligible_activity_model_name(report_key):
            continue
        if model_key:
            eligible_keys.add(model_key)
        else:
            eligible_legacy_names.add(model_name)
    if not eligible_keys and not eligible_legacy_names:
        return queryset.none()
    predicate = Q()
    if eligible_keys:
        predicate |= Q(model_key__in=eligible_keys)
    if eligible_legacy_names:
        predicate |= Q(model_key__isnull=True, model_name__in=eligible_legacy_names)
    return queryset.filter(predicate)


def _format_actions(counter, strings):
    return [
        {
            "key": key,
            "label": strings.get(f"action_{str(key or '').lower()}", key or strings.get("user_report_unknown")),
            "count": count,
        }
        for key, count in counter.most_common()
    ]


def _format_model_actions(model_action_map, strings):
    models = []
    for model_key, action_counter in model_action_map.items():
        models.append({
            "key": model_key,
            "label": translate_activity_log_model_name(model_key, strings=strings),
            "count": sum(action_counter.values()),
            "actions": _format_actions(action_counter, strings),
        })
    models.sort(key=lambda item: item["count"], reverse=True)
    return models


def build_activity_windows(activity_qs, *, strings=None):
    strings = strings or get_strings()
    now = timezone.now()
    month_start, _ = get_report_window_bounds("month", now=now)
    week_start, _ = get_report_window_bounds("week", now=now)
    window_action = {"week": Counter(), "month": Counter(), "all": Counter()}
    window_model_action = {
        "week": defaultdict(Counter),
        "month": defaultdict(Counter),
        "all": defaultdict(Counter),
    }
    for action, row_model_key, row_model_name, created_at in activity_qs.values_list(
        "action", "model_key", "model_name", "created_at"
    ):
        report_key = activity_report_key(row_model_key, row_model_name)
        if not is_report_eligible_activity_model_name(report_key):
            continue
        group_key = report_key or strings.get("user_report_unknown")
        for window in ("all",):
            window_action[window][action] += 1
            window_model_action[window][group_key][action] += 1
        if created_at and month_start and created_at >= month_start:
            window_action["month"][action] += 1
            window_model_action["month"][group_key][action] += 1
            if week_start and created_at >= week_start:
                window_action["week"][action] += 1
                window_model_action["week"][group_key][action] += 1
    return {
        window: {
            "activity_count": sum(window_action[window].values()),
            "models": _format_model_actions(window_model_action[window], strings),
            "action_counts": _format_actions(window_action[window], strings),
        }
        for window in ("week", "month", "all")
    }


def _visible_user_queryset(actor):
    User = apps.get_model(settings.AUTH_USER_MODEL)
    qs = User._default_manager.select_related("profile__scope").order_by("username")
    if getattr(actor, "is_superuser", False) or is_global_staff(actor) or not is_scope_enabled():
        return qs
    if is_central_staff(actor):
        return qs.filter(profile__scope__isnull=True)
    actor_scope = get_user_scope(actor)
    if actor_scope is not None:
        return qs.filter(profile__scope=actor_scope)
    return qs.none()


def _aggregation_queryset(queryset):
    """Strip ordering/joins before grouped count queries."""
    return queryset.select_related(None).order_by()


def _count_grouped(queryset, *fields):
    return (
        _aggregation_queryset(queryset)
        .values(*fields)
        .annotate(count=Count("pk"))
        .order_by("-count")
    )


def _overview_cache_scope(actor):
    scope = get_user_scope(actor)
    return {
        "user_id": getattr(actor, "pk", None),
        "scope_enabled": is_scope_enabled(),
        "scope_id": getattr(scope, "pk", None),
        "superuser": bool(getattr(actor, "is_superuser", False)),
        "global": bool(is_global_staff(actor)),
        "central": bool(is_central_staff(actor)),
    }


def _overview_stats_cache_key(actor, window, filters):
    payload = {
        "schema": REPORT_OVERVIEW_CACHE_SCHEMA_VERSION,
        "scope": _overview_cache_scope(actor),
        "window": window,
        "filters": {
            "q": filters.get("q") or "",
            "custom_start": filters.get("custom_start") or "",
            "custom_end": filters.get("custom_end") or "",
            "models": sorted(filters.get("models") or []),
            "operations": sorted(filters.get("operations") or []),
        },
        "language": get_current_language_code(),
        "config": repr(_reports_config()),
    }
    raw = json.dumps(payload, sort_keys=True, default=str, separators=(",", ":"))
    import hashlib
    return "dlux:reports:overview:%s" % hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _safe_cache_get(key):
    try:
        return cache.get(key)
    except Exception:
        logger.warning("Report overview cache read failed", exc_info=True)
        return None


def _safe_cache_set(key, value, timeout):
    try:
        cache.set(key, value, timeout=timeout)
    except Exception:
        logger.warning("Report overview cache write failed", exc_info=True)


def _build_reports_overview_stats(current_qs, previous_qs, all_qs, *, strings):
    unknown = strings.get("user_report_unknown")

    user_counts = Counter()
    for row in _count_grouped(current_qs, "created_by__username"):
        user_counts[row["created_by__username"] or unknown] += row["count"]

    model_counts = Counter()
    for row in _count_grouped(current_qs, "model_key", "model_name"):
        key = activity_report_key(row["model_key"], row["model_name"]) or unknown
        model_counts[key] += row["count"]

    action_counts = Counter()
    for row in _count_grouped(current_qs, "action"):
        action_counts[row["action"] or unknown] += row["count"]

    day_counts = []
    day_rows = (
        _aggregation_queryset(current_qs)
        .annotate(report_day=TruncDate("created_at", tzinfo=timezone.get_current_timezone()))
        .values("report_day")
        .annotate(count=Count("pk"))
        .order_by("-report_day")
    )
    for row in day_rows:
        day = row["report_day"]
        label = day.isoformat() if day else unknown
        day_counts.append({"label": label, "count": row["count"]})

    current_total = sum(user_counts.values())
    active_days = len(day_counts)
    active_users = len(user_counts)
    previous_period_total = previous_qs.count() if previous_qs is not None else 0
    all_total = all_qs.count()

    return {
        "current_total": current_total,
        "previous_period_total": previous_period_total,
        "previous_week_total": previous_period_total,
        "all_total": all_total,
        "delta": current_total - previous_period_total,
        "average_per_active_day": round(current_total / active_days, 2) if active_days else 0,
        "average_per_user": round(current_total / active_users, 2) if active_users else 0,
        "users": [
            {"label": key, "count": count}
            for key, count in user_counts.most_common()
        ],
        "models": [
            {"key": key, "label": translate_activity_log_model_name(key, strings=strings), "count": count}
            for key, count in model_counts.most_common()
        ],
        "actions": _format_actions(action_counts, strings),
        "days": day_counts,
    }


def _filter_values(filters, plural_key, legacy_key):
    value = filters.get(plural_key)
    if value is None:
        value = filters.get(legacy_key)
    if value is None:
        return None
    if isinstance(value, str):
        value = [value]
    return [str(item).strip() for item in value if str(item or "").strip()]


def _report_filter_catalog(activity_qs, *, strings):
    catalog = {}
    for model in get_report_eligible_models():
        key = model._meta.label_lower
        catalog[key] = {
            "key": key,
            "label": str(model._meta.verbose_name),
            "registered": True,
        }
    rows = (
        _aggregation_queryset(activity_qs)
        .values_list("model_key", "model_name")
        .distinct()
    )
    for model_key, model_name in rows:
        activity_key = activity_report_key(model_key, model_name)
        if not activity_key or not is_report_eligible_activity_model_name(activity_key):
            continue
        resolved_model = resolve_model_by_name(model_name) if not model_key else None
        key = resolved_model._meta.label_lower if resolved_model is not None else activity_key
        entry = catalog.setdefault(key, {"key": key, "registered": False})
        entry["label"] = translate_activity_log_model_name(key, strings=strings)
        entry["model_key"] = model_key or ""
        if not model_key and model_name:
            legacy_names = entry.setdefault("legacy_names", [])
            if model_name not in legacy_names:
                legacy_names.append(model_name)
    models_catalog = sorted(catalog.values(), key=lambda item: item["label"].casefold())
    operation_keys = list(
        _aggregation_queryset(activity_qs)
        .exclude(action="")
        .order_by("action")
        .values_list("action", flat=True)
        .distinct()
    )
    operations_catalog = [
        {
            "key": key,
            "label": strings.get(f"action_{str(key).lower()}", key),
        }
        for key in operation_keys
    ]
    return models_catalog, operations_catalog


def _apply_report_builder_filters(queryset, *, model_catalog, selected_models, selected_operations):
    if selected_models is not None:
        if not selected_models:
            return queryset.none()
        selected = set(selected_models)
        predicate = Q()
        for entry in model_catalog:
            if entry["key"] not in selected:
                continue
            model_key = entry.get("model_key") or (entry["key"] if "." in entry["key"] else "")
            if model_key:
                predicate |= Q(model_key=model_key)
            legacy_names = entry.get("legacy_names") or []
            if legacy_names:
                predicate |= Q(model_key__isnull=True, model_name__in=legacy_names)
                predicate |= Q(model_key="", model_name__in=legacy_names)
        if not predicate:
            return queryset.none()
        queryset = queryset.filter(predicate)
    if selected_operations is not None:
        if not selected_operations:
            return queryset.none()
        queryset = queryset.filter(action__in=selected_operations)
    return queryset


def _report_period_details(window, custom_start, custom_end, *, now=None):
    now = timezone.localtime(now or timezone.now())
    errors = []
    parsed_start = _parse_report_date(custom_start)
    parsed_end = _parse_report_date(custom_end)
    if window == "custom":
        if parsed_start is None:
            errors.append("start")
        if parsed_end is None:
            errors.append("end")
        if parsed_start and parsed_end and parsed_end < parsed_start:
            errors.append("order")
        effective_start = parsed_start or now.date()
        effective_end = parsed_end or effective_start
        if effective_end < effective_start:
            effective_start, effective_end = effective_end, effective_start
        custom_start = effective_start.isoformat()
        custom_end = effective_end.isoformat()
    else:
        custom_start = ""
        custom_end = ""
    start, end = get_report_window_bounds(
        window,
        now=now,
        custom_start=custom_start,
        custom_end=custom_end,
    )
    start_label = timezone.localtime(start).date().isoformat() if start else ""
    if end is not None:
        # Custom ranges carry an exclusive upper bound (midnight after the last
        # day), so the inclusive label is the day before it.
        end_label = (timezone.localtime(end).date() - timedelta(days=1)).isoformat()
    elif start is not None:
        # Week/month/quarter/half-year/year are deliberately open-ended: they run
        # from their start through now. Label that end as today so "Active period"
        # reads as the range it filters on rather than a lone start date.
        end_label = now.date().isoformat()
    else:
        end_label = ""
    # En dash rather than an arrow: it carries no direction, so the range reads
    # correctly in both LTR and RTL.
    range_label = f"{start_label} – {end_label}" if end_label and end_label != start_label else start_label
    return {
        "custom_start": custom_start,
        "custom_end": custom_end,
        "start": start,
        "end": end,
        "start_label": start_label,
        "end_label": end_label,
        "range_label": range_label,
        "errors": errors,
    }


def build_reports_overview(actor, *, window="week", filters=None, now=None):
    strings = get_strings()
    window = normalize_report_window(window)
    filters = filters or {}
    now = now or timezone.now()
    period = _report_period_details(
        window,
        filters.get("custom_start"),
        filters.get("custom_end"),
        now=now,
    )
    base_qs = filter_report_eligible_activity(
        apply_report_scope(_base_activity_queryset(), actor)
    )
    model_catalog, operation_catalog = _report_filter_catalog(base_qs, strings=strings)
    model_keys = {item["key"] for item in model_catalog}
    operation_keys = {item["key"] for item in operation_catalog}
    requested_models = _filter_values(filters, "models", "model")
    requested_operations = _filter_values(filters, "operations", "action")
    explicit_builder = str(filters.get("builder") or "") == "1"
    if requested_models is None and not explicit_builder:
        selected_models = sorted(model_keys)
    else:
        selected_models = [key for key in (requested_models or []) if key in model_keys]
    if requested_operations is None and not explicit_builder:
        selected_operations = sorted(operation_keys)
    else:
        selected_operations = [key for key in (requested_operations or []) if key in operation_keys]

    filtered_qs = _apply_report_builder_filters(
        base_qs,
        model_catalog=model_catalog,
        selected_models=selected_models,
        selected_operations=selected_operations,
    )

    keyword = str(filters.get("q") or "").strip()
    if keyword:
        filtered_qs = filtered_qs.filter(
            Q(created_by__username__icontains=keyword)
            | Q(created_by__first_name__icontains=keyword)
            | Q(created_by__last_name__icontains=keyword)
            | Q(action__icontains=keyword)
            | Q(model_name__icontains=keyword)
            | Q(number__icontains=keyword)
        )
    all_qs = filtered_qs
    current_qs = apply_report_window(
        filtered_qs,
        window,
        now=now,
        custom_start=period["custom_start"],
        custom_end=period["custom_end"],
    ).order_by("-created_at")
    previous_start, previous_end = get_previous_period_bounds(
        window,
        now=now,
        custom_start=period["custom_start"],
        custom_end=period["custom_end"],
    )
    previous_qs = None
    if previous_start is not None:
        previous_qs = filtered_qs.filter(
            created_at__gte=previous_start,
            created_at__lt=previous_end,
        )

    users = _visible_user_queryset(actor)
    stats = None
    cache_seconds = get_reports_overview_cache_seconds()
    if cache_seconds:
        cache_key = _overview_stats_cache_key(
            actor,
            window,
            {
                "q": keyword,
                "custom_start": period["custom_start"],
                "custom_end": period["custom_end"],
                "models": selected_models,
                "operations": selected_operations,
            },
        )
        stats = _safe_cache_get(cache_key)
    else:
        cache_key = None
    if stats is None:
        stats = _build_reports_overview_stats(current_qs, previous_qs, all_qs, strings=strings)
        if cache_seconds and cache_key:
            _safe_cache_set(cache_key, stats, cache_seconds)

    criteria = {
        "window": window,
        "q": keyword,
        "custom_start": period["custom_start"],
        "custom_end": period["custom_end"],
        "models": selected_models,
        "operations": selected_operations,
        "builder": "1",
    }
    selected_model_set = set(selected_models)
    selected_operation_set = set(selected_operations)
    for item in model_catalog:
        item["selected"] = item["key"] in selected_model_set
    for item in operation_catalog:
        item["selected"] = item["key"] in selected_operation_set
    return {
        "window": window,
        "filters": criteria,
        "criteria": criteria,
        "period": period,
        "current_total": stats["current_total"],
        "previous_period_total": stats["previous_period_total"],
        "previous_week_total": stats["previous_week_total"],
        "all_total": stats["all_total"],
        "delta": stats["delta"],
        "average_per_active_day": stats["average_per_active_day"],
        "average_per_user": stats["average_per_user"],
        "users": stats["users"],
        "models": stats["models"],
        "actions": stats["actions"],
        "days": stats["days"],
        "visible_users": users,
        "recent_activity": list(current_qs[:50]),
        "available_models": model_catalog,
        "available_actions": operation_catalog,
        "selected_model_count": len(selected_models),
        "selected_operation_count": len(selected_operations),
        "activity_qs": current_qs,
    }


def _chart_series(items, *, limit=REPORT_CHART_TOP_N, other_label=""):
    """Top-N label/count pairs, with the remainder folded into one 'Other' slice."""
    series = [
        {"label": str(item.get("label") or item.get("key") or ""), "count": int(item.get("count") or 0)}
        for item in items
    ]
    if limit and len(series) > limit:
        head = series[:limit]
        remainder = sum(item["count"] for item in series[limit:])
        if remainder:
            head.append({"label": other_label, "count": remainder})
        return head
    return series


def build_report_chart_data(overview, *, strings=None):
    """Chart-ready series for the printable report.

    ``days`` is re-ordered chronologically (the stats query returns newest-first)
    so the trend chart reads left-to-right. The ranked breakdowns are capped at
    ``REPORT_CHART_TOP_N`` so a project with hundreds of models still prints a
    legible chart; the operation mix — the only chart that colours by identity —
    is capped tighter, at the eight validated categorical hue slots.
    """
    strings = strings or get_strings()
    other_label = strings.get("reports_print_other", "Other")
    return {
        "days": [
            {"label": str(item.get("label") or ""), "count": int(item.get("count") or 0)}
            for item in reversed(overview.get("days") or [])
        ],
        "models": _chart_series(overview.get("models") or [], other_label=other_label),
        "actions": _chart_series(
            overview.get("actions") or [],
            limit=REPORT_CHART_CATEGORICAL_TOP_N,
            other_label=other_label,
        ),
        "users": _chart_series(overview.get("users") or [], other_label=other_label),
    }


def _scope_model_queryset(model, actor):
    manager = getattr(model, "all_objects", model._default_manager)
    qs = manager.all()
    if hasattr(model, "deleted_at"):
        qs = qs.filter(deleted_at__isnull=True)
    if not is_scope_enabled():
        return qs
    field_names = {field.name for field in model._meta.get_fields()}
    if "scope" not in field_names:
        return qs
    if getattr(actor, "is_superuser", False) or is_global_staff(actor):
        return qs
    if is_central_staff(actor):
        return qs.filter(scope__isnull=True)
    actor_scope = get_user_scope(actor)
    if actor_scope is not None:
        return qs.filter(scope=actor_scope)
    return qs.none()


_BACKUP_TIMESTAMP_CANDIDATES = ("created_at", "created", "created_on", "date_created", "timestamp")
_BACKUP_LABEL_FIELD_CANDIDATES = (
    "number",
    "document_number",
    "reference_number",
    "registration_number",
    "serial_number",
    "code",
    "name",
    "title",
)


def get_backup_storage_prefix():
    """Storage-relative directory where generated backup zips are kept.

    Lives under the default storage (usually MEDIA_ROOT) so web and worker
    containers can share it; deployments must block direct HTTP access to it
    (e.g. an nginx `deny all` on /media/<prefix>/) — downloads always go
    through the permission-checked Django view.
    """
    prefix = str(_reports_config().get("backup_storage_prefix") or "dlux_backups").strip("/")
    return prefix or "dlux_backups"


def _backup_label_field(model):
    configured = _reports_config().get("backup_label_fields", {})
    field_name = ""
    if isinstance(configured, dict):
        field_name = str(configured.get(model._meta.label_lower) or "").strip()
    candidates = (field_name,) if field_name else _BACKUP_LABEL_FIELD_CANDIDATES
    for candidate in candidates:
        if not candidate:
            continue
        try:
            field = model._meta.get_field(candidate)
        except Exception:
            continue
        if getattr(field, "concrete", False) and not getattr(field, "is_relation", False):
            return field.name
    return ""


def _model_natural_key_fields(model):
    """Best-effort field list composing a model's natural key, or ``[]``.

    System backups serialize foreign keys with ``use_natural_foreign_keys=True``,
    so an FK to a model that defines ``natural_key()`` is stored as that natural
    key (e.g. ``["alice"]``) rather than a PK. Only confidently-derivable cases
    are reported — currently models exposing ``USERNAME_FIELD`` (the auth user) —
    so the viewer can map such references to a label. When this returns ``[]``
    the viewer falls back to displaying the raw natural key, which is itself
    already human-readable.
    """
    if not callable(getattr(model, "natural_key", None)):
        return []
    username_field = getattr(model, "USERNAME_FIELD", None)
    if username_field:
        try:
            model._meta.get_field(username_field)
        except Exception:
            return []
        return [username_field]
    return []


def build_relation_schema(models_list):
    """Per-model relation + label metadata baked into the backup manifest.

    Lets the standalone ``.dlb`` viewer resolve FK/O2O/M2M references to readable
    labels without the originating project. Shape::

        {"app.model": {"label_field": "name",
                       "natural_key_fields": ["username"],   # omitted when empty
                       "relations": {"field": {"kind": "fk|o2o|m2m",
                                               "to": "app.target"}}}}

    Only relations that ``dumpdata`` actually serializes are recorded (concrete
    FK/O2O fields and M2M fields backed by an auto-created through table), keyed
    by the same name the serializer uses in each record's ``fields`` map.
    Defensive: any per-model failure is skipped rather than breaking the backup,
    since this is a viewer convenience, not restore data.
    """
    schema = {}
    for model in models_list:
        try:
            meta = model._meta
            relations = {}
            for field in meta.concrete_fields:
                if not field.is_relation or field.related_model is None:
                    continue
                kind = "o2o" if field.one_to_one else "fk" if field.many_to_one else ""
                if not kind:
                    continue
                relations[field.name] = {
                    "kind": kind,
                    "to": field.related_model._meta.label_lower,
                }
            for field in meta.local_many_to_many:
                if field.related_model is None:
                    continue
                # Only auto-created through tables are serialized by dumpdata.
                if not field.remote_field.through._meta.auto_created:
                    continue
                relations[field.name] = {
                    "kind": "m2m",
                    "to": field.related_model._meta.label_lower,
                }
            label_field = _backup_label_field(model)
            if not label_field:
                # User-like models rarely have a name/title field but read well
                # by their login identifier (e.g. username).
                username_field = getattr(model, "USERNAME_FIELD", None)
                if username_field:
                    try:
                        meta.get_field(username_field)
                        label_field = username_field
                    except Exception:
                        pass
            entry = {
                "label_field": label_field,
                "relations": relations,
            }
            natural_key_fields = _model_natural_key_fields(model)
            if natural_key_fields:
                entry["natural_key_fields"] = natural_key_fields
            schema[meta.label_lower] = entry
        except Exception:
            continue
    return schema


def _safe_archive_segment(value, *, max_length=80):
    value = unicodedata.normalize("NFKC", str(value or "")).strip()
    value = re.sub(r"[\\/\x00-\x1f\x7f]+", "-", value)
    value = re.sub(r"\s+", "-", value)
    value = re.sub(r"-+", "-", value).strip(" .-")
    return value[:max_length].rstrip(" .-")


def backup_record_folder(record, *, label_field=None):
    """Human-readable base folder for one report-backup record."""
    field_name = _backup_label_field(record.__class__) if label_field is None else label_field
    label = _safe_archive_segment(getattr(record, field_name, "")) if field_name else ""
    if label:
        field_label = _safe_archive_segment(field_name.replace("_", "-"), max_length=30)
        return f"{field_label}-{label}"
    object_label = _safe_archive_segment(str(record))
    if object_label and object_label != _safe_archive_segment(record.pk, max_length=80):
        return f"record-{object_label}"
    return "record"


def _backup_timestamp_field(model):
    """Field used to window-filter a model's rows, or None to include all rows."""
    config_map = _reports_config().get("backup_window_fields", {})
    if isinstance(config_map, dict) and model._meta.label_lower in config_map:
        configured = str(config_map.get(model._meta.label_lower) or "").strip()
        return configured or None
    names = {
        field.name
        for field in model._meta.get_fields()
        if isinstance(field, (models.DateTimeField, models.DateField))
    }
    for candidate in _BACKUP_TIMESTAMP_CANDIDATES:
        if candidate in names:
            return candidate
    return None


def _backup_model_queryset(model, actor, window, *, custom_start=None, custom_end=None):
    qs = _scope_model_queryset(model, actor)
    start, end = get_report_window_bounds(
        window,
        custom_start=custom_start,
        custom_end=custom_end,
    )
    if start is None and end is None:
        return qs
    field = _backup_timestamp_field(model)
    if not field:
        # No timestamp to filter on: include everything rather than silently dropping rows.
        return qs
    if start is not None:
        qs = qs.filter(**{f"{field}__gte": start})
    if end is not None:
        qs = qs.filter(**{f"{field}__lt": end})
    return qs


def _iter_queryset_by_pk(qs, chunk_size=200):
    """Yield model rows in bounded primary-key pages without server-side cursors."""
    try:
        chunk_size = int(chunk_size)
    except (TypeError, ValueError):
        chunk_size = 200
    chunk_size = max(1, chunk_size)
    pk_attname = qs.model._meta.pk.attname
    ordered = qs.order_by(pk_attname)
    last_pk = None
    while True:
        page = ordered
        if last_pk is not None:
            page = page.filter(**{f"{pk_attname}__gt": last_pk})
        batch = list(page[:chunk_size])
        if not batch:
            break
        for obj in batch:
            yield obj
        last_pk = getattr(batch[-1], pk_attname)


class _CursorlessJSONSerializer(JsonSerializer):
    """Backup serializer variant that avoids QuerySet.iterator() entirely."""

    def handle_m2m_field(self, obj, field):
        if not field.remote_field.through._meta.auto_created:
            return
        related = getattr(obj, field.name)
        if self.use_natural_foreign_keys and hasattr(
            field.remote_field.model, "natural_key"
        ):
            self._current[field.name] = [value.natural_key() for value in related.all()]
            return
        related_qs = related.select_related(None).only("pk")
        self._current[field.name] = [
            self._value_from_field(value, value._meta.pk)
            for value in related_qs
        ]


def stream_model_into_zip(
    zf,
    model,
    qs,
    manifest,
    *,
    serialize_kwargs=None,
    object_transform=None,
    human_record_folders=False,
    include_files=True,
    include_records=True,
    step_callback=None,
):
    """Stream one model's records (JSON) and its file-field contents into an
    open backup ZipFile, recording everything in ``manifest``.

    Shared by the reports backup and the full system backup: the serializer
    writes straight into the zip entry and storage files are chunk-copied, so
    peak memory stays flat no matter how many records/PDFs are covered.

    ``include_files=False`` produces a data-only export: record JSON is still
    written (FileField *names* are preserved in the data), but the referenced
    media blobs are not copied into the archive. A restore then leaves existing
    media on disk untouched (``_restore_files`` only touches files the manifest
    lists). This is what the inline updater's pre-update backup uses so a quick
    code/schema update is not gated on copying gigabytes of unchanged uploads.

    ``include_records=False`` is the mirror image: copy the media and record the
    manifest entries, but write no ``data/`` JSON. The periodic reports ZIP uses
    it because that archive is a deliverable read by people — its data ships as
    the entries workbook — and is never restored. The restorable ``.dlb`` always
    keeps the JSON.

    ``step_callback(stage, done, total)`` — ``stage`` being ``"rows"`` or
    ``"files"`` — reports movement *inside* one model. Without it a model holding
    thousands of uploads looks frozen for as long as it takes to copy them, which
    is indistinguishable from a dead worker.
    """
    meta = model._meta
    model_key = meta.label_lower
    total_rows = qs.count()
    manifest["models"].append({"model": model_key, "count": total_rows})

    def report(stage, done, total):
        if step_callback:
            step_callback(stage, done, total)

    def serialized_objects():
        for position, obj in enumerate(_iter_queryset_by_pk(qs, chunk_size=200), start=1):
            report("rows", position, total_rows)
            yield object_transform(obj) if object_transform else obj

    if include_records:
        with zf.open(f"data/{meta.app_label}/{meta.model_name}.json", mode="w") as raw_stream:
            text_stream = io.TextIOWrapper(raw_stream, encoding="utf-8")
            serializer = _CursorlessJSONSerializer()
            serializer.serialize(
                serialized_objects(),
                stream=text_stream,
                **(serialize_kwargs or {}),
            )
            text_stream.flush()
            text_stream.detach()
    if not include_files:
        return
    file_fields = [
        field for field in meta.get_fields()
        if isinstance(field, (models.FileField, models.ImageField))
    ]
    if not file_fields:
        return
    record_label_field = _backup_label_field(model) if human_record_folders else ""
    folder_counts = {}
    for position, record in enumerate(_iter_queryset_by_pk(qs, chunk_size=100), start=1):
        report("files", position, total_rows)
        if human_record_folders:
            base_folder = backup_record_folder(record, label_field=record_label_field)
            folder_counts[base_folder] = folder_counts.get(base_folder, 0) + 1
            occurrence = folder_counts[base_folder]
            record_folder = base_folder if occurrence == 1 else f"{base_folder}--{occurrence}"
        else:
            record_folder = str(record.pk)
        for field in file_fields:
            file_value = getattr(record, field.name, None)
            if not file_value or not getattr(file_value, "name", ""):
                continue
            archive_name = f"files/{meta.app_label}/{meta.model_name}/{record_folder}/{field.name}/{file_value.name.split('/')[-1]}"
            try:
                with default_storage.open(file_value.name, "rb") as fh, \
                        zf.open(archive_name, mode="w") as dest:
                    shutil.copyfileobj(fh, dest, 256 * 1024)
                manifest["files"].append({
                    "model": model_key,
                    "pk": record.pk,
                    "record_folder": record_folder,
                    "record_label_field": record_label_field,
                    "field": field.name,
                    "path": archive_name,
                    # Original storage name — what a restore writes the file back as.
                    "name": file_value.name,
                })
            except Exception:
                manifest["missing_files"].append({
                    "model": model_key,
                    "pk": record.pk,
                    "field": field.name,
                    "name": file_value.name,
                })


def _models_for_report_criteria(criteria):
    selected = criteria.get("models")
    if selected is None:
        return get_report_eligible_models()
    selected_keys = {
        normalize_activity_log_model_key(value)
        for value in selected
        if str(value or "").strip()
    }
    return [
        model for model in get_report_eligible_models()
        if _model_keys(model) & selected_keys
    ]


# ── Model entry export (XLSX) ───────────────────────────────────────────────
#
# The builder's XLSX download is a *data* export: one sheet per selected model
# holding that model's real rows for the chosen period. The analytical figures
# live in the printable report instead, where charts and layout can carry them.

_XLSX_SHEET_TITLE_INVALID = re.compile(r"[\[\]:*?/\\]")
# Control characters openpyxl refuses to write into a shared string.
_XLSX_ILLEGAL_CHARS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")
_XLSX_MAX_CELL_LENGTH = 32767
_XLSX_MAX_COLUMN_WIDTH = 60
# Never place credential material in a spreadsheet a user can mail around. Matched
# against the concrete field name, so a project's custom user/token models are
# covered without needing configuration.
_EXPORT_SENSITIVE_FIELD_PARTS = (
    "password",
    "passphrase",
    "secret",
    "token",
    "api_key",
    "apikey",
    "private_key",
    "session_key",
    "recovery_code",
    "backup_code",
    "salt",
)


def get_report_entries_row_limit():
    """Per-model row cap for the XLSX entry export.

    Bounds worst-case workbook size/memory on a project with very large tables.
    Override with DLUX_CONFIG['reports']['entries_row_limit'].
    """
    try:
        value = int(_reports_config().get("entries_row_limit", REPORT_ENTRIES_ROW_LIMIT))
    except (TypeError, ValueError):
        value = REPORT_ENTRIES_ROW_LIMIT
    return max(1, min(value, 1000000))


def _unique_sheet_title(title, used):
    """Excel-legal, <=31 char, case-insensitively unique worksheet title."""
    base = _XLSX_SHEET_TITLE_INVALID.sub(" ", str(title or "")).strip()
    base = re.sub(r"\s+", " ", base)[:31].strip() or "Sheet"
    candidate = base
    index = 2
    while candidate.casefold() in used:
        suffix = f"~{index}"
        candidate = f"{base[:31 - len(suffix)].strip()}{suffix}"
        index += 1
    used.add(candidate.casefold())
    return candidate


def _is_sensitive_export_field(field):
    name = str(getattr(field, "name", "") or "").lower()
    return any(part in name for part in _EXPORT_SENSITIVE_FIELD_PARTS)


def _configured_export_exclusions(model):
    configured = _reports_config().get("entries_exclude_fields", {})
    if not isinstance(configured, dict):
        return set()
    values = configured.get(model._meta.label_lower, [])
    if isinstance(values, str):
        values = [values]
    return {str(value).strip() for value in values or [] if str(value or "").strip()}


def get_report_entry_fields(model):
    """Concrete fields exported for one model, minus credential-bearing ones."""
    excluded = _configured_export_exclusions(model)
    return [
        field for field in model._meta.concrete_fields
        if field.name not in excluded and not _is_sensitive_export_field(field)
    ]


def _xlsx_text(value):
    return _XLSX_ILLEGAL_CHARS.sub("", str(value))[:_XLSX_MAX_CELL_LENGTH]


def _xlsx_cell_value(value):
    """Coerce a model field value into something openpyxl can write."""
    if value is None:
        return ""
    if isinstance(value, (bool, int, float, Decimal)):
        return value
    if isinstance(value, datetime):
        # openpyxl cannot write tz-aware datetimes; localize then drop the tzinfo.
        return timezone.localtime(value).replace(tzinfo=None) if timezone.is_aware(value) else value
    if isinstance(value, (date, time)):
        return value
    if isinstance(value, (bytes, bytearray, memoryview)):
        return _xlsx_text(f"<{len(bytes(value))} bytes>")
    if isinstance(value, (list, tuple, set, dict)):
        return _xlsx_text(json.dumps(value, ensure_ascii=False, default=str, sort_keys=True))
    return _xlsx_text(value)


def _export_field_value(record, field):
    if getattr(field, "choices", None):
        display = getattr(record, f"get_{field.name}_display", None)
        if callable(display):
            return display()
    if field.is_relation:
        related = getattr(record, field.name, None)
        return str(related) if related is not None else ""
    value = getattr(record, field.attname, None)
    if isinstance(field, (models.FileField, models.ImageField)):
        return getattr(value, "name", "") or ""
    return value


def _entry_export_queryset(model, actor, criteria, overview):
    """Rows exported for one model under the builder's period/scope selection."""
    if model._meta.label_lower == "dlux.activitylog" and overview is not None:
        # Reuse the already-filtered activity queryset so the operation selection
        # applies here exactly as it does on screen and in the ZIP.
        return overview["activity_qs"]
    qs = _backup_model_queryset(
        model,
        actor,
        normalize_report_window(criteria.get("window")),
        custom_start=criteria.get("custom_start"),
        custom_end=criteria.get("custom_end"),
    )
    related = [
        field.name for field in model._meta.concrete_fields
        if field.is_relation and field.related_model is not None
    ]
    return qs.select_related(*related) if related else qs


def build_model_entries_xlsx(actor, overview):
    """Workbook of the *actual entries* of every model the builder selected.

    One sheet per model with that model's own columns and rows, filtered by the
    selected period and the caller's scope — not the report's aggregate figures,
    which are the printable report's job.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    strings = get_strings()
    criteria = overview["criteria"]
    period = overview["period"]
    row_limit = get_report_entries_row_limit()
    used_titles = set()

    wb = Workbook()
    info = wb.active
    info.title = _unique_sheet_title(strings.get("reports_sheet_export_info", "Export Info"), used_titles)

    index_rows = []
    for model in _models_for_report_criteria(criteria):
        fields = get_report_entry_fields(model)
        if not fields:
            continue
        title = _unique_sheet_title(str(model._meta.verbose_name_plural), used_titles)
        ws = wb.create_sheet(title)
        headers = [_xlsx_text(field.verbose_name or field.name) for field in fields]
        ws.append(headers)
        widths = [len(header) for header in headers]

        exported = 0
        qs = _entry_export_queryset(model, actor, criteria, overview)
        for record in _iter_queryset_by_pk(qs, chunk_size=200):
            if exported >= row_limit:
                break
            row = [_xlsx_cell_value(_export_field_value(record, field)) for field in fields]
            ws.append(row)
            exported += 1
            for column, value in enumerate(row):
                widths[column] = max(widths[column], len(str(value or "")))

        total = qs.count()
        if not exported:
            ws.append([strings.get("reports_export_empty_model", "No entries in the selected period.")])
        for cell in ws[1]:
            cell.font = Font(bold=True)
        ws.freeze_panes = "A2"
        for column, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(column)].width = min(max(width + 2, 12), _XLSX_MAX_COLUMN_WIDTH)
        index_rows.append([
            str(model._meta.verbose_name_plural),
            model._meta.label_lower,
            title,
            exported,
            strings.get("reports_export_truncated_yes", "Yes") if total > exported
            else strings.get("reports_export_truncated_no", "No"),
        ])

    info_rows = [
        [strings.get("user_report_field"), strings.get("user_report_value")],
        [strings.get("reports_export_generated_at", "Generated at"),
         timezone.localtime().replace(tzinfo=None)],
        [strings.get("reports_window"), criteria["window"]],
        [strings.get("reports_custom_start", "Start date"), period["start_label"]],
        [strings.get("reports_custom_end", "End date"), period["end_label"]],
        [strings.get("reports_models_included", "Models included"), len(index_rows)],
        [strings.get("reports_export_row_limit", "Row limit per model"), row_limit],
        [],
    ]
    if index_rows:
        info_rows.append([
            strings.get("user_report_model"),
            strings.get("reports_export_model_key", "Model key"),
            strings.get("reports_export_sheet", "Sheet"),
            strings.get("reports_export_rows", "Rows exported"),
            strings.get("reports_export_truncated", "Truncated"),
        ])
        info_rows.extend(index_rows)
    else:
        info_rows.append([
            strings.get("reports_export_no_models", "No models were selected, so no entry sheets were produced."),
        ])
    for row in info_rows:
        info.append(row)
    for cell in info[1]:
        cell.font = Font(bold=True)
    if index_rows:
        for cell in info[len(info_rows) - len(index_rows)]:
            cell.font = Font(bold=True)
    for column in info.columns:
        max_len = max((len(str(cell.value or "")) for cell in column), default=12)
        info.column_dimensions[column[0].column_letter].width = min(max(max_len + 2, 14), _XLSX_MAX_COLUMN_WIDTH)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream.getvalue()


REPORT_ZIP_WORKBOOK_NAME = "entries.xlsx"


def write_backup_zip(actor, fileobj, *, window="all", criteria=None, progress_callback=None):
    """Stream a scope-aware report package into ``fileobj``.

    The archive is a periodic deliverable for people, not a restore artifact
    (that is the system ``.dlb``), so it carries exactly two things: the entries
    workbook the Export Entries button produces, and the media those records
    reference, foldered by business identifier. No serialized JSON — the same
    normalized period/model/operation selection drives both parts, so the
    browser, the spreadsheet, and the ZIP always agree.
    """
    window = normalize_backup_window(window)
    requested_filters = dict(criteria or {})
    if criteria:
        requested_filters["builder"] = "1"
    overview = build_reports_overview(actor, window=window, filters=requested_filters)
    criteria = overview["criteria"]
    manifest = {
        "generated_at": timezone.now().isoformat(),
        "window": window,
        "selection": criteria,
        "models": [],
        "files": [],
        "missing_files": [],
        "report_artifacts": [REPORT_ZIP_WORKBOOK_NAME],
    }
    models_to_export = _models_for_report_criteria(criteria)
    total_models = max(len(models_to_export), 1)
    strings = get_strings()
    with zipfile.ZipFile(fileobj, "w", zipfile.ZIP_DEFLATED) as zf:
        for index, model in enumerate(models_to_export):
            if progress_callback:
                progress_callback(
                    5 + int((index / total_models) * 85),
                    strings.get("backup_progress_model", "Backing up {model}...").format(
                        model=str(model._meta.verbose_name),
                    ),
                )
            stream_model_into_zip(
                zf,
                model,
                _entry_export_queryset(model, actor, criteria, overview),
                manifest,
                human_record_folders=True,
                include_records=False,
            )
            if progress_callback:
                progress_callback(
                    5 + int(((index + 1) / total_models) * 85),
                    strings.get("backup_progress_model_done", "Backed up {model}.").format(
                        model=str(model._meta.verbose_name),
                    ),
                )
        if progress_callback:
            progress_callback(92, strings.get("reports_backup_building_workbook", "Building report workbook..."))
        zf.writestr(REPORT_ZIP_WORKBOOK_NAME, build_model_entries_xlsx(actor, overview))
        zf.writestr("manifest.json", json.dumps(manifest, indent=2, ensure_ascii=False))
    return manifest


def build_backup_zip(request, window="all", criteria=None):
    """In-memory backup build kept for backward compatibility (small datasets only)."""
    buffer = BytesIO()
    manifest = write_backup_zip(request.user, buffer, window=window, criteria=criteria)
    buffer.seek(0)
    log_user_action(
        request,
        "EXPORT",
        model_name="Dlux Reports Backup",
        details={
            "window": manifest["window"],
            "models": len(manifest["models"]),
            "files": len(manifest["files"]),
        },
    )
    return buffer.getvalue(), manifest


# ── Background backup runs (Celery when available, synchronous fallback) ────


def _report_backup_model():
    return apps.get_model("dlux", "ReportBackup")


def report_backup_celery_available():
    """True when the backup task can actually run in the background right now:
    celery importable, broker reachable, and at least one live worker."""
    if not _reports_config().get("backup_use_celery", True):
        return False
    try:
        from .tasks import build_report_backup_task
    except Exception:
        return False
    if build_report_backup_task is None:
        return False
    try:
        app = build_report_backup_task.app
        with app.connection_for_write() as conn:
            conn.ensure_connection(max_retries=0, timeout=2)
        return bool(app.control.ping(timeout=1.0))
    except Exception:
        return False


def dispatch_report_backup(backup):
    """Queue the backup build on Celery. Returns True when queued."""
    if not report_backup_celery_available():
        return False
    from .tasks import build_report_backup_task
    try:
        build_report_backup_task.apply_async(args=[backup.pk], retry=False)
        return True
    except Exception:
        logger.exception("Failed to queue report backup pk=%s", backup.pk)
        return False


def _prune_report_backups(user, keep=3):
    ReportBackup = _report_backup_model()
    stale = list(
        ReportBackup.objects.filter(user=user, status=ReportBackup.STATUS_COMPLETED)
        .order_by("-created_at")[keep:]
    ) + list(
        ReportBackup.objects.filter(
            user=user,
            status=ReportBackup.STATUS_FAILED,
            created_at__lt=timezone.now() - timedelta(days=7),
        )
    )
    for old in stale:
        if old.file_path:
            try:
                default_storage.delete(old.file_path)
            except Exception:
                pass
        old.delete()


def run_report_backup(backup_pk):
    """Build the zip for a ReportBackup row and store it under the backup prefix.

    Runs inside the Celery worker (or inline as a last resort). Status/result
    are persisted on the row so the web process can poll over the shared DB.
    """
    ReportBackup = _report_backup_model()
    backup = ReportBackup.objects.filter(pk=backup_pk).first()
    if backup is None or backup.status not in (ReportBackup.STATUS_PENDING,):
        return backup
    logger.info(
        "Starting report backup pk=%s token=%s window=%s user_id=%s",
        backup.pk,
        backup.token,
        backup.window,
        backup.user_id,
    )
    backup.status = ReportBackup.STATUS_RUNNING
    backup.started_at = timezone.now()
    backup.save(update_fields=["status", "started_at"])
    from .backup_progress import finish_backup_progress, set_backup_progress, start_backup_progress
    start_backup_progress(backup)
    set_backup_progress(backup, 2, get_strings().get("backup_progress_preparing", "Preparing backup..."))
    try:
        with tempfile.TemporaryFile() as tmp:
            manifest = write_backup_zip(
                backup.user,
                tmp,
                window=backup.window,
                criteria=backup.criteria or None,
                progress_callback=lambda percent, message: set_backup_progress(backup, percent, message),
            )
            size = tmp.tell()
            tmp.seek(0)
            set_backup_progress(backup, 95, get_strings().get("backup_progress_storing", "Storing backup artifact..."))
            saved_path = default_storage.save(
                f"{get_backup_storage_prefix()}/{backup.token}.zip",
                File(tmp),
            )
        backup.file_path = saved_path
        backup.file_size = size
        backup.model_count = len(manifest["models"])
        backup.file_count = len(manifest["files"])
        backup.missing_file_count = len(manifest["missing_files"])
        backup.status = ReportBackup.STATUS_COMPLETED
        backup.completed_at = timezone.now()
        backup.error = ""
        backup.save()
        finish_backup_progress(backup, success=True)
        logger.info(
            "Completed report backup pk=%s token=%s size=%s models=%s files=%s missing=%s",
            backup.pk,
            backup.token,
            backup.file_size,
            backup.model_count,
            backup.file_count,
            backup.missing_file_count,
        )
        UserActivityLog = apps.get_model("dlux", "ActivityLog")
        UserActivityLog.safe_log(
            user=backup.user,
            action="EXPORT",
            model_name="Dlux Reports Backup",
            details={
                "window": backup.window,
                "criteria": backup.criteria,
                "models": backup.model_count,
                "files": backup.file_count,
            },
        )
        _prune_report_backups(backup.user)
    except Exception as exc:
        logger.exception("Report backup pk=%s failed", backup_pk)
        backup.status = ReportBackup.STATUS_FAILED
        backup.completed_at = timezone.now()
        backup.error = str(exc)[:1000]
        backup.save(update_fields=["status", "completed_at", "error"])
        finish_backup_progress(backup, success=False, error=backup.error)
    return backup
