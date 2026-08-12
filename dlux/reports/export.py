"""XLSX entry export."""

import json
import re
from datetime import date, datetime, time
from decimal import Decimal
from io import BytesIO
from django.db import models
from django.utils import timezone
from ..translations import get_strings
from ..utils import normalize_activity_log_model_key

from .config import _reports_config, get_report_entries_row_limit
from .eligibility import _model_keys, get_report_eligible_models
from .queries import _backup_model_queryset, _iter_queryset_by_pk
from .windows import normalize_report_window


def _models_for_report_criteria(criteria):
    selected = criteria.get("models")
    if selected is None:
        return get_report_eligible_models()
    selected_keys = {
        normalize_activity_log_model_key(value)
        for value in selected
        if str(value or "").strip()
    }
    return [
        model for model in get_report_eligible_models()
        if _model_keys(model) & selected_keys
    ]


_XLSX_SHEET_TITLE_INVALID = re.compile(r"[\[\]:*?/\\]")


_XLSX_ILLEGAL_CHARS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


_XLSX_MAX_CELL_LENGTH = 32767


_XLSX_MAX_COLUMN_WIDTH = 60


_EXPORT_SENSITIVE_FIELD_PARTS = (
    "password",
    "passphrase",
    "secret",
    "token",
    "api_key",
    "apikey",
    "private_key",
    "session_key",
    "recovery_code",
    "backup_code",
    "salt",
)


def _unique_sheet_title(title, used):
    """Excel-legal, <=31 char, case-insensitively unique worksheet title."""
    base = _XLSX_SHEET_TITLE_INVALID.sub(" ", str(title or "")).strip()
    base = re.sub(r"\s+", " ", base)[:31].strip() or "Sheet"
    candidate = base
    index = 2
    while candidate.casefold() in used:
        suffix = f"~{index}"
        candidate = f"{base[:31 - len(suffix)].strip()}{suffix}"
        index += 1
    used.add(candidate.casefold())
    return candidate


def _is_sensitive_export_field(field):
    name = str(getattr(field, "name", "") or "").lower()
    return any(part in name for part in _EXPORT_SENSITIVE_FIELD_PARTS)


def _configured_export_exclusions(model):
    configured = _reports_config().get("entries_exclude_fields", {})
    if not isinstance(configured, dict):
        return set()
    values = configured.get(model._meta.label_lower, [])
    if isinstance(values, str):
        values = [values]
    return {str(value).strip() for value in values or [] if str(value or "").strip()}


def get_report_entry_fields(model):
    """Concrete fields exported for one model, minus credential-bearing ones."""
    excluded = _configured_export_exclusions(model)
    return [
        field for field in model._meta.concrete_fields
        if field.name not in excluded and not _is_sensitive_export_field(field)
    ]


def _xlsx_text(value):
    return _XLSX_ILLEGAL_CHARS.sub("", str(value))[:_XLSX_MAX_CELL_LENGTH]


def _xlsx_cell_value(value):
    """Coerce a model field value into something openpyxl can write."""
    if value is None:
        return ""
    if isinstance(value, (bool, int, float, Decimal)):
        return value
    if isinstance(value, datetime):
        # openpyxl cannot write tz-aware datetimes; localize then drop the tzinfo.
        return timezone.localtime(value).replace(tzinfo=None) if timezone.is_aware(value) else value
    if isinstance(value, (date, time)):
        return value
    if isinstance(value, (bytes, bytearray, memoryview)):
        return _xlsx_text(f"<{len(bytes(value))} bytes>")
    if isinstance(value, (list, tuple, set, dict)):
        return _xlsx_text(json.dumps(value, ensure_ascii=False, default=str, sort_keys=True))
    return _xlsx_text(value)


def _export_field_value(record, field):
    if getattr(field, "choices", None):
        display = getattr(record, f"get_{field.name}_display", None)
        if callable(display):
            return display()
    if field.is_relation:
        related = getattr(record, field.name, None)
        return str(related) if related is not None else ""
    value = getattr(record, field.attname, None)
    if isinstance(field, (models.FileField, models.ImageField)):
        return getattr(value, "name", "") or ""
    return value


def _entry_export_queryset(model, actor, criteria, overview):
    """Rows exported for one model under the builder's period/scope selection."""
    if model._meta.label_lower == "dlux.activitylog" and overview is not None:
        # Reuse the already-filtered activity queryset so the operation selection
        # applies here exactly as it does on screen and in the ZIP.
        return overview["activity_qs"]
    qs = _backup_model_queryset(
        model,
        actor,
        normalize_report_window(criteria.get("window")),
        custom_start=criteria.get("custom_start"),
        custom_end=criteria.get("custom_end"),
    )
    related = [
        field.name for field in model._meta.concrete_fields
        if field.is_relation and field.related_model is not None
    ]
    return qs.select_related(*related) if related else qs


def build_model_entries_xlsx(actor, overview):
    """Workbook of the *actual entries* of every model the builder selected.

    One sheet per model with that model's own columns and rows, filtered by the
    selected period and the caller's scope — not the report's aggregate figures,
    which are the printable report's job.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    strings = get_strings()
    criteria = overview["criteria"]
    period = overview["period"]
    row_limit = get_report_entries_row_limit()
    used_titles = set()

    wb = Workbook()
    info = wb.active
    info.title = _unique_sheet_title(strings.get("reports_sheet_export_info", "Export Info"), used_titles)

    index_rows = []
    for model in _models_for_report_criteria(criteria):
        fields = get_report_entry_fields(model)
        if not fields:
            continue
        title = _unique_sheet_title(str(model._meta.verbose_name_plural), used_titles)
        ws = wb.create_sheet(title)
        headers = [_xlsx_text(field.verbose_name or field.name) for field in fields]
        ws.append(headers)
        widths = [len(header) for header in headers]

        exported = 0
        qs = _entry_export_queryset(model, actor, criteria, overview)
        for record in _iter_queryset_by_pk(qs, chunk_size=200):
            if exported >= row_limit:
                break
            row = [_xlsx_cell_value(_export_field_value(record, field)) for field in fields]
            ws.append(row)
            exported += 1
            for column, value in enumerate(row):
                widths[column] = max(widths[column], len(str(value or "")))

        total = qs.count()
        if not exported:
            ws.append([strings.get("reports_export_empty_model", "No entries in the selected period.")])
        for cell in ws[1]:
            cell.font = Font(bold=True)
        ws.freeze_panes = "A2"
        for column, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(column)].width = min(max(width + 2, 12), _XLSX_MAX_COLUMN_WIDTH)
        index_rows.append([
            str(model._meta.verbose_name_plural),
            model._meta.label_lower,
            title,
            exported,
            strings.get("reports_export_truncated_yes", "Yes") if total > exported
            else strings.get("reports_export_truncated_no", "No"),
        ])

    info_rows = [
        [strings.get("user_report_field"), strings.get("user_report_value")],
        [strings.get("reports_export_generated_at", "Generated at"),
         timezone.localtime().replace(tzinfo=None)],
        [strings.get("reports_window"), criteria["window"]],
        [strings.get("reports_custom_start", "Start date"), period["start_label"]],
        [strings.get("reports_custom_end", "End date"), period["end_label"]],
        [strings.get("reports_models_included", "Models included"), len(index_rows)],
        [strings.get("reports_export_row_limit", "Row limit per model"), row_limit],
        [],
    ]
    if index_rows:
        info_rows.append([
            strings.get("user_report_model"),
            strings.get("reports_export_model_key", "Model key"),
            strings.get("reports_export_sheet", "Sheet"),
            strings.get("reports_export_rows", "Rows exported"),
            strings.get("reports_export_truncated", "Truncated"),
        ])
        info_rows.extend(index_rows)
    else:
        info_rows.append([
            strings.get("reports_export_no_models", "No models were selected, so no entry sheets were produced."),
        ])
    for row in info_rows:
        info.append(row)
    for cell in info[1]:
        cell.font = Font(bold=True)
    if index_rows:
        for cell in info[len(info_rows) - len(index_rows)]:
            cell.font = Font(bold=True)
    for column in info.columns:
        max_len = max((len(str(cell.value or "")) for cell in column), default=12)
        info.column_dimensions[column[0].column_letter].width = min(max(max_len + 2, 14), _XLSX_MAX_COLUMN_WIDTH)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream.getvalue()
