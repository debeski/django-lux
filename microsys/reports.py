import io
import json
import logging
import shutil
import tempfile
import zipfile
from collections import Counter, defaultdict
from datetime import datetime, time, timedelta
from io import BytesIO

from django.apps import apps
from django.conf import settings
from django.core.cache import cache
from django.core import serializers
from django.core.files import File
from django.core.files.storage import default_storage
from django.db import models
from django.db.models import Count, Q
from django.db.models.functions import TruncDate
from django.utils import timezone

from .translations import get_current_language_code, get_strings
from .utils import (
    get_user_scope,
    is_central_staff,
    is_global_staff,
    is_scope_enabled,
    log_user_action,
    normalize_activity_log_model_key,
    resolve_model_by_name,
    translate_activity_log_model_name,
)


logger = logging.getLogger("microsys")

REPORT_WINDOWS = {"week", "month", "all"}
_MICROSYS_REPORT_EXCLUDED_KEYS = {
    "auth",
    "authentication",
    "session",
    "sessions",
    "systemsettings",
    "system_settings",
    "scope",
    "scopesettings",
    "scope_settings",
    "profile",
    "user_profile",
    "trusteddevice",
    "trusted_device",
    "userknowndevice",
    "known_device",
    "known_devices",
    "userpresencesession",
    "presence_session",
    "presence_sessions",
    "publicregistration",
    "public_registration",
    "preferences",
    "password",
    "totp",
    "otp",
    "backup_code",
    "backup_codes",
    # Microsys-internal meta actions (not project work; classified as system interactions)
    "microsys_reports_backup",
    "reports_backup",
}

# Verbose names of the high-frequency operational tracking models that are no longer
# logged (see signals.EXCLUDED_MODELS). Historical rows may still exist; use
# exclude_log_noise() to keep them out of activity feeds and reports entirely.
LOG_NOISE_MODEL_NAMES = ("Trusted Device", "Known Device", "Presence Session")


def exclude_log_noise(queryset):
    """Drop operational tracking-model rows (presence/device churn) from a log queryset."""
    return queryset.exclude(model_name__in=LOG_NOISE_MODEL_NAMES)
_EXCLUDED_APP_LABELS = {
    "admin",
    "auth",
    "contenttypes",
    "sessions",
    "messages",
    "staticfiles",
}


def normalize_report_window(value):
    value = str(value or "week").strip().lower()
    return value if value in REPORT_WINDOWS else "week"


def normalize_backup_window(value):
    """Like normalize_report_window but defaults to 'all' (the historical backup scope)."""
    value = str(value or "all").strip().lower()
    return value if value in REPORT_WINDOWS else "all"


def _reports_config():
    config = getattr(settings, "MICROSYS_CONFIG", {}).get("reports", {})
    return config if isinstance(config, dict) else {}


def get_reports_overview_cache_seconds():
    """Optional read-through cache TTL for the expensive report overview stats.

    Defaults to 0 to keep the overview exact out of the box. Deployments with a
    shared Django cache/Redis can set MICROSYS_CONFIG['reports']['overview_cache_seconds']
    to a small value to absorb repeated dashboard loads.
    """
    try:
        value = int(_reports_config().get("overview_cache_seconds", 0))
    except (TypeError, ValueError):
        value = 0
    return max(0, min(value, 3600))


def _normalized_config_set(key):
    values = _reports_config().get(key, [])
    if isinstance(values, str):
        values = [values]
    return {normalize_activity_log_model_key(value) for value in values or [] if str(value or "").strip()}


def _model_keys(model):
    meta = model._meta
    values = {
        meta.label_lower,
        f"{meta.app_label}.{meta.model_name}",
        meta.model_name,
        meta.object_name,
        str(meta.verbose_name),
        str(meta.verbose_name_plural),
    }
    return {normalize_activity_log_model_key(value) for value in values if str(value or "").strip()}


def _model_is_section(model):
    marker = getattr(model, "is_section", None)
    if isinstance(marker, bool):
        return marker
    if marker is not None:
        return True
    return bool(getattr(model._meta, "is_section", False))


def is_report_eligible_model(model):
    if model is None:
        return False
    meta = model._meta
    if meta.abstract or not meta.managed:
        return False
    keys = _model_keys(model)
    if keys & _normalized_config_set("exclude_models"):
        return False
    if keys & _normalized_config_set("include_models"):
        return True
    if meta.app_label in _EXCLUDED_APP_LABELS:
        return False
    if meta.app_label == "microsys":
        return _model_is_section(model)
    return True


def is_report_eligible_activity_model_name(model_name):
    raw = str(model_name or "").strip()
    if not raw:
        return False
    # Non-ASCII verbose names (e.g. translated/Arabic model labels stored by the
    # activity logger) normalize to an empty key. Don't reject them outright — the
    # key-based checks below simply won't match, and eligibility falls through to
    # model resolution so the decision is made on the underlying model instead of
    # on whether its label happens to be ASCII.
    normalized = normalize_activity_log_model_key(raw)
    if normalized and normalized in _MICROSYS_REPORT_EXCLUDED_KEYS:
        return False
    if normalized in _normalized_config_set("exclude_activity"):
        return False
    if normalized in _normalized_config_set("include_activity"):
        return True
    model = resolve_model_by_name(raw)
    if model is not None:
        return is_report_eligible_model(model)
    return True


def get_report_eligible_models():
    return [model for model in apps.get_models() if is_report_eligible_model(model)]


def _week_start_index():
    value = _reports_config().get("week_start", 0)
    try:
        value = int(value)
    except (TypeError, ValueError):
        value = 0
    return value if 0 <= value <= 6 else 0


def get_report_window_bounds(window, *, now=None):
    window = normalize_report_window(window)
    if window == "all":
        return None, None
    now = timezone.localtime(now or timezone.now())
    if window == "month":
        start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        return start, None
    week_start = _week_start_index()
    start_date = now.date() - timedelta(days=(now.weekday() - week_start) % 7)
    start = timezone.make_aware(datetime.combine(start_date, time.min), timezone.get_current_timezone())
    return start, None


def get_previous_week_bounds(*, now=None):
    current_start, _ = get_report_window_bounds("week", now=now)
    previous_start = current_start - timedelta(days=7)
    return previous_start, current_start


def _base_activity_queryset():
    ActivityLog = apps.get_model("microsys", "UserActivityLog")
    manager = getattr(ActivityLog, "all_objects", ActivityLog._default_manager)
    qs = manager.filter(deleted_at__isnull=True) if hasattr(ActivityLog, "deleted_at") else manager.all()
    return qs.select_related("created_by", "created_by__profile__scope", "scope")


def apply_report_scope(queryset, actor):
    if not is_scope_enabled():
        return queryset
    if getattr(actor, "is_superuser", False) or is_global_staff(actor):
        return queryset
    if is_central_staff(actor):
        return queryset.filter(scope__isnull=True)
    actor_scope = get_user_scope(actor)
    if actor_scope is not None:
        return queryset.filter(scope=actor_scope)
    return queryset.none()


def apply_report_window(queryset, window, *, now=None):
    start, end = get_report_window_bounds(window, now=now)
    if start is not None:
        queryset = queryset.filter(created_at__gte=start)
    if end is not None:
        queryset = queryset.filter(created_at__lt=end)
    return queryset


def get_visible_report_activity(actor, *, window="all", now=None):
    qs = apply_report_scope(_base_activity_queryset(), actor)
    qs = apply_report_window(qs, window, now=now)
    return qs.order_by("-created_at")


def activity_report_key(model_key, model_name):
    """Locale-independent identity for an activity row.

    Prefer the stable ``app_label.model_name`` key; fall back to the (possibly
    translated) display label for legacy rows and non-model events. Grouping and
    eligibility must run on this, never on the raw ``model_name`` label, which
    changes with the active language.
    """
    key = str(model_key or "").strip()
    if key:
        return key
    return str(model_name or "").strip()


def log_report_key(log):
    """``activity_report_key`` for a UserActivityLog instance."""
    return activity_report_key(getattr(log, "model_key", None), getattr(log, "model_name", None))


def filter_report_eligible_activity(queryset):
    eligible_keys = set()
    eligible_legacy_names = set()
    identity_rows = queryset.select_related(None).order_by().values_list("model_key", "model_name").distinct()
    for model_key, model_name in identity_rows:
        report_key = activity_report_key(model_key, model_name)
        if not report_key or not is_report_eligible_activity_model_name(report_key):
            continue
        if model_key:
            eligible_keys.add(model_key)
        else:
            eligible_legacy_names.add(model_name)
    if not eligible_keys and not eligible_legacy_names:
        return queryset.none()
    predicate = Q()
    if eligible_keys:
        predicate |= Q(model_key__in=eligible_keys)
    if eligible_legacy_names:
        predicate |= Q(model_key__isnull=True, model_name__in=eligible_legacy_names)
    return queryset.filter(predicate)


def _format_actions(counter, strings):
    return [
        {
            "key": key,
            "label": strings.get(f"action_{str(key or '').lower()}", key or strings.get("user_report_unknown")),
            "count": count,
        }
        for key, count in counter.most_common()
    ]


def _format_model_actions(model_action_map, strings):
    models = []
    for model_key, action_counter in model_action_map.items():
        models.append({
            "key": model_key,
            "label": translate_activity_log_model_name(model_key, strings=strings),
            "count": sum(action_counter.values()),
            "actions": _format_actions(action_counter, strings),
        })
    models.sort(key=lambda item: item["count"], reverse=True)
    return models


def build_activity_windows(activity_qs, *, strings=None):
    strings = strings or get_strings()
    now = timezone.now()
    month_start, _ = get_report_window_bounds("month", now=now)
    week_start, _ = get_report_window_bounds("week", now=now)
    window_action = {"week": Counter(), "month": Counter(), "all": Counter()}
    window_model_action = {
        "week": defaultdict(Counter),
        "month": defaultdict(Counter),
        "all": defaultdict(Counter),
    }
    for action, row_model_key, row_model_name, created_at in activity_qs.values_list(
        "action", "model_key", "model_name", "created_at"
    ):
        report_key = activity_report_key(row_model_key, row_model_name)
        if not is_report_eligible_activity_model_name(report_key):
            continue
        group_key = report_key or strings.get("user_report_unknown")
        for window in ("all",):
            window_action[window][action] += 1
            window_model_action[window][group_key][action] += 1
        if created_at and month_start and created_at >= month_start:
            window_action["month"][action] += 1
            window_model_action["month"][group_key][action] += 1
            if week_start and created_at >= week_start:
                window_action["week"][action] += 1
                window_model_action["week"][group_key][action] += 1
    return {
        window: {
            "activity_count": sum(window_action[window].values()),
            "models": _format_model_actions(window_model_action[window], strings),
            "action_counts": _format_actions(window_action[window], strings),
        }
        for window in ("week", "month", "all")
    }


def _visible_user_queryset(actor):
    User = apps.get_model(settings.AUTH_USER_MODEL)
    qs = User._default_manager.select_related("profile__scope").order_by("username")
    if getattr(actor, "is_superuser", False) or is_global_staff(actor) or not is_scope_enabled():
        return qs
    if is_central_staff(actor):
        return qs.filter(profile__scope__isnull=True)
    actor_scope = get_user_scope(actor)
    if actor_scope is not None:
        return qs.filter(profile__scope=actor_scope)
    return qs.none()


def _aggregation_queryset(queryset):
    """Strip ordering/joins before grouped count queries."""
    return queryset.select_related(None).order_by()


def _count_grouped(queryset, *fields):
    return (
        _aggregation_queryset(queryset)
        .values(*fields)
        .annotate(count=Count("pk"))
        .order_by("-count")
    )


def _overview_cache_scope(actor):
    scope = get_user_scope(actor)
    return {
        "user_id": getattr(actor, "pk", None),
        "scope_enabled": is_scope_enabled(),
        "scope_id": getattr(scope, "pk", None),
        "superuser": bool(getattr(actor, "is_superuser", False)),
        "global": bool(is_global_staff(actor)),
        "central": bool(is_central_staff(actor)),
    }


def _overview_stats_cache_key(actor, window, filters):
    payload = {
        "scope": _overview_cache_scope(actor),
        "window": window,
        "filters": {
            "q": filters.get("q") or "",
            "model": filters.get("model") or "",
            "action": filters.get("action") or "",
        },
        "language": get_current_language_code(),
        "config": repr(_reports_config()),
    }
    raw = json.dumps(payload, sort_keys=True, default=str, separators=(",", ":"))
    import hashlib
    return "microsys:reports:overview:%s" % hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _safe_cache_get(key):
    try:
        return cache.get(key)
    except Exception:
        logger.warning("Report overview cache read failed", exc_info=True)
        return None


def _safe_cache_set(key, value, timeout):
    try:
        cache.set(key, value, timeout=timeout)
    except Exception:
        logger.warning("Report overview cache write failed", exc_info=True)


def _build_reports_overview_stats(current_qs, previous_qs, all_qs, *, strings):
    unknown = strings.get("user_report_unknown")

    user_counts = Counter()
    for row in _count_grouped(current_qs, "created_by__username"):
        user_counts[row["created_by__username"] or unknown] += row["count"]

    model_counts = Counter()
    for row in _count_grouped(current_qs, "model_key", "model_name"):
        key = activity_report_key(row["model_key"], row["model_name"]) or unknown
        model_counts[key] += row["count"]

    action_counts = Counter()
    for row in _count_grouped(current_qs, "action"):
        action_counts[row["action"] or unknown] += row["count"]

    day_counts = []
    day_rows = (
        _aggregation_queryset(current_qs)
        .annotate(report_day=TruncDate("created_at", tzinfo=timezone.get_current_timezone()))
        .values("report_day")
        .annotate(count=Count("pk"))
        .order_by("-report_day")
    )
    for row in day_rows:
        day = row["report_day"]
        label = day.isoformat() if day else unknown
        day_counts.append({"label": label, "count": row["count"]})

    current_total = sum(user_counts.values())
    active_days = len(day_counts)
    active_users = len(user_counts)
    previous_week_total = previous_qs.count()
    all_total = all_qs.count()

    return {
        "current_total": current_total,
        "previous_week_total": previous_week_total,
        "all_total": all_total,
        "delta": current_total - previous_week_total,
        "average_per_active_day": round(current_total / active_days, 2) if active_days else 0,
        "average_per_user": round(current_total / active_users, 2) if active_users else 0,
        "users": [
            {"label": key, "count": count}
            for key, count in user_counts.most_common()
        ],
        "models": [
            {"key": key, "label": translate_activity_log_model_name(key, strings=strings), "count": count}
            for key, count in model_counts.most_common()
        ],
        "actions": _format_actions(action_counts, strings),
        "days": day_counts,
        "available_models": list(
            _aggregation_queryset(current_qs)
            .order_by("model_name")
            .values_list("model_name", flat=True)
            .distinct()
        ),
        "available_actions": list(
            _aggregation_queryset(current_qs)
            .order_by("action")
            .values_list("action", flat=True)
            .distinct()
        ),
    }


def build_reports_overview(actor, *, window="week", filters=None):
    strings = get_strings()
    window = normalize_report_window(window)
    filters = filters or {}
    current_qs = filter_report_eligible_activity(get_visible_report_activity(actor, window=window))
    previous_start, previous_end = get_previous_week_bounds()
    previous_qs = filter_report_eligible_activity(
        apply_report_scope(_base_activity_queryset(), actor).filter(
            created_at__gte=previous_start,
            created_at__lt=previous_end,
        )
    )
    all_qs = filter_report_eligible_activity(get_visible_report_activity(actor, window="all"))

    keyword = str(filters.get("q") or "").strip()
    if keyword:
        current_qs = current_qs.filter(
            Q(created_by__username__icontains=keyword)
            | Q(created_by__first_name__icontains=keyword)
            | Q(created_by__last_name__icontains=keyword)
            | Q(action__icontains=keyword)
            | Q(model_name__icontains=keyword)
            | Q(number__icontains=keyword)
        )
    model_filter = str(filters.get("model") or "").strip()
    if model_filter:
        current_qs = current_qs.filter(model_name=model_filter)
    action_filter = str(filters.get("action") or "").strip()
    if action_filter:
        current_qs = current_qs.filter(action=action_filter)

    users = _visible_user_queryset(actor)
    stats = None
    cache_seconds = get_reports_overview_cache_seconds()
    if cache_seconds:
        cache_key = _overview_stats_cache_key(
            actor,
            window,
            {"q": keyword, "model": model_filter, "action": action_filter},
        )
        stats = _safe_cache_get(cache_key)
    else:
        cache_key = None
    if stats is None:
        stats = _build_reports_overview_stats(current_qs, previous_qs, all_qs, strings=strings)
        if cache_seconds and cache_key:
            _safe_cache_set(cache_key, stats, cache_seconds)

    return {
        "window": window,
        "filters": {"q": keyword, "model": model_filter, "action": action_filter},
        "current_total": stats["current_total"],
        "previous_week_total": stats["previous_week_total"],
        "all_total": stats["all_total"],
        "delta": stats["delta"],
        "average_per_active_day": stats["average_per_active_day"],
        "average_per_user": stats["average_per_user"],
        "users": stats["users"],
        "models": stats["models"],
        "actions": stats["actions"],
        "days": stats["days"],
        "visible_users": users,
        "recent_activity": list(current_qs[:50]),
        "available_models": stats["available_models"],
        "available_actions": stats["available_actions"],
        "activity_qs": current_qs,
    }


def build_reports_overview_xlsx(overview):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    strings = get_strings()
    wb = Workbook()

    def write_rows(ws, rows):
        for row in rows:
            ws.append(row)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for column in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in column)
            ws.column_dimensions[column[0].column_letter].width = min(max(max_len + 2, 12), 60)

    summary = wb.active
    summary.title = strings.get("reports_sheet_summary", "Summary")[:31]
    write_rows(summary, [
        [strings.get("user_report_field"), strings.get("user_report_value")],
        [strings.get("reports_window"), overview["window"]],
        [strings.get("reports_current_total"), overview["current_total"]],
        [strings.get("reports_previous_week_total"), overview["previous_week_total"]],
        [strings.get("reports_all_total"), overview["all_total"]],
        [strings.get("reports_delta"), overview["delta"]],
        [strings.get("reports_average_per_day"), overview["average_per_active_day"]],
        [strings.get("reports_average_per_user"), overview["average_per_user"]],
    ])

    for sheet_key, title_key, rows in (
        ("users", "reports_by_user", overview["users"]),
        ("models", "reports_by_model", overview["models"]),
        ("actions", "reports_by_action", overview["actions"]),
        ("days", "reports_by_day", overview["days"]),
    ):
        ws = wb.create_sheet(strings.get(title_key, sheet_key.title())[:31])
        write_rows(ws, [[strings.get("user_report_value"), strings.get("user_report_count")]] + [
            [item.get("label") or item.get("key"), item.get("count")] for item in rows
        ])

    logs = wb.create_sheet(strings.get("user_report_sheet_logs", "Logs")[:31])
    log_rows = [[
        strings.get("user_report_timestamp"),
        strings.get("user_report_username"),
        strings.get("user_report_action"),
        strings.get("user_report_model"),
        strings.get("user_report_count"),
    ]]
    for log in overview["activity_qs"][:1000]:
        created_at = timezone.localtime(log.created_at).replace(tzinfo=None) if log.created_at else None
        log_rows.append([
            created_at,
            log.created_by.get_username() if log.created_by else "",
            log.action,
            translate_activity_log_model_name(log.model_name, strings=strings),
            log.number or "",
        ])
    write_rows(logs, log_rows)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream.getvalue()


def _scope_model_queryset(model, actor):
    manager = getattr(model, "all_objects", model._default_manager)
    qs = manager.all()
    if hasattr(model, "deleted_at"):
        qs = qs.filter(deleted_at__isnull=True)
    if not is_scope_enabled():
        return qs
    field_names = {field.name for field in model._meta.get_fields()}
    if "scope" not in field_names:
        return qs
    if getattr(actor, "is_superuser", False) or is_global_staff(actor):
        return qs
    if is_central_staff(actor):
        return qs.filter(scope__isnull=True)
    actor_scope = get_user_scope(actor)
    if actor_scope is not None:
        return qs.filter(scope=actor_scope)
    return qs.none()


_BACKUP_TIMESTAMP_CANDIDATES = ("created_at", "created", "created_on", "date_created", "timestamp")


def get_backup_storage_prefix():
    """Storage-relative directory where generated backup zips are kept.

    Lives under the default storage (usually MEDIA_ROOT) so web and worker
    containers can share it; deployments must block direct HTTP access to it
    (e.g. an nginx `deny all` on /media/<prefix>/) — downloads always go
    through the permission-checked Django view.
    """
    prefix = str(_reports_config().get("backup_storage_prefix") or "microsys_backups").strip("/")
    return prefix or "microsys_backups"


def _backup_timestamp_field(model):
    """Field used to window-filter a model's rows, or None to include all rows."""
    config_map = _reports_config().get("backup_window_fields", {})
    if isinstance(config_map, dict) and model._meta.label_lower in config_map:
        configured = str(config_map.get(model._meta.label_lower) or "").strip()
        return configured or None
    names = {
        field.name
        for field in model._meta.get_fields()
        if isinstance(field, (models.DateTimeField, models.DateField))
    }
    for candidate in _BACKUP_TIMESTAMP_CANDIDATES:
        if candidate in names:
            return candidate
    return None


def _backup_model_queryset(model, actor, window):
    qs = _scope_model_queryset(model, actor)
    start, end = get_report_window_bounds(window)
    if start is None and end is None:
        return qs
    field = _backup_timestamp_field(model)
    if not field:
        # No timestamp to filter on: include everything rather than silently dropping rows.
        return qs
    if start is not None:
        qs = qs.filter(**{f"{field}__gte": start})
    if end is not None:
        qs = qs.filter(**{f"{field}__lt": end})
    return qs


def stream_model_into_zip(zf, model, qs, manifest, *, serialize_kwargs=None, object_transform=None):
    """Stream one model's records (JSON) and its file-field contents into an
    open backup ZipFile, recording everything in ``manifest``.

    Shared by the reports backup and the full system backup: the serializer
    writes straight into the zip entry and storage files are chunk-copied, so
    peak memory stays flat no matter how many records/PDFs are covered.
    """
    meta = model._meta
    model_key = meta.label_lower
    manifest["models"].append({"model": model_key, "count": qs.count()})
    def serialized_objects():
        for obj in qs.iterator(chunk_size=200):
            yield object_transform(obj) if object_transform else obj

    with zf.open(f"data/{meta.app_label}/{meta.model_name}.json", mode="w") as raw_stream:
        text_stream = io.TextIOWrapper(raw_stream, encoding="utf-8")
        serializers.serialize(
            "json",
            serialized_objects(),
            stream=text_stream,
            **(serialize_kwargs or {}),
        )
        text_stream.flush()
        text_stream.detach()
    file_fields = [
        field for field in meta.get_fields()
        if isinstance(field, (models.FileField, models.ImageField))
    ]
    if not file_fields:
        return
    for record in qs.iterator(chunk_size=100):
        for field in file_fields:
            file_value = getattr(record, field.name, None)
            if not file_value or not getattr(file_value, "name", ""):
                continue
            archive_name = f"files/{meta.app_label}/{meta.model_name}/{record.pk}/{field.name}/{file_value.name.split('/')[-1]}"
            try:
                with default_storage.open(file_value.name, "rb") as fh, \
                        zf.open(archive_name, mode="w") as dest:
                    shutil.copyfileobj(fh, dest, 256 * 1024)
                manifest["files"].append({
                    "model": model_key,
                    "pk": record.pk,
                    "field": field.name,
                    "path": archive_name,
                    # Original storage name — what a restore writes the file back as.
                    "name": file_value.name,
                })
            except Exception:
                manifest["missing_files"].append({
                    "model": model_key,
                    "pk": record.pk,
                    "field": field.name,
                    "name": file_value.name,
                })


def write_backup_zip(actor, fileobj, *, window="all"):
    """Stream a scope-aware, window-filtered backup zip into ``fileobj``."""
    window = normalize_backup_window(window)
    manifest = {
        "generated_at": timezone.now().isoformat(),
        "window": window,
        "models": [],
        "files": [],
        "missing_files": [],
    }
    with zipfile.ZipFile(fileobj, "w", zipfile.ZIP_DEFLATED) as zf:
        for model in get_report_eligible_models():
            qs = _backup_model_queryset(model, actor, window)
            stream_model_into_zip(zf, model, qs, manifest)
        zf.writestr("manifest.json", json.dumps(manifest, indent=2, ensure_ascii=False))
    return manifest


def build_backup_zip(request, window="all"):
    """In-memory backup build kept for backward compatibility (small datasets only)."""
    buffer = BytesIO()
    manifest = write_backup_zip(request.user, buffer, window=window)
    buffer.seek(0)
    log_user_action(
        request,
        "EXPORT",
        model_name="Microsys Reports Backup",
        details={
            "window": manifest["window"],
            "models": len(manifest["models"]),
            "files": len(manifest["files"]),
        },
    )
    return buffer.getvalue(), manifest


# ── Background backup runs (Celery when available, synchronous fallback) ────


def _report_backup_model():
    return apps.get_model("microsys", "ReportBackup")


def report_backup_celery_available():
    """True when the backup task can actually run in the background right now:
    celery importable, broker reachable, and at least one live worker."""
    if not _reports_config().get("backup_use_celery", True):
        return False
    try:
        from .tasks import build_report_backup_task
    except Exception:
        return False
    if build_report_backup_task is None:
        return False
    try:
        app = build_report_backup_task.app
        with app.connection_for_write() as conn:
            conn.ensure_connection(max_retries=0, timeout=2)
        return bool(app.control.ping(timeout=1.0))
    except Exception:
        return False


def dispatch_report_backup(backup):
    """Queue the backup build on Celery. Returns True when queued."""
    if not report_backup_celery_available():
        return False
    from .tasks import build_report_backup_task
    try:
        build_report_backup_task.apply_async(args=[backup.pk], retry=False)
        return True
    except Exception:
        logger.exception("Failed to queue report backup pk=%s", backup.pk)
        return False


def _prune_report_backups(user, keep=3):
    ReportBackup = _report_backup_model()
    stale = list(
        ReportBackup.objects.filter(user=user, status=ReportBackup.STATUS_COMPLETED)
        .order_by("-created_at")[keep:]
    ) + list(
        ReportBackup.objects.filter(
            user=user,
            status=ReportBackup.STATUS_FAILED,
            created_at__lt=timezone.now() - timedelta(days=7),
        )
    )
    for old in stale:
        if old.file_path:
            try:
                default_storage.delete(old.file_path)
            except Exception:
                pass
        old.delete()


def run_report_backup(backup_pk):
    """Build the zip for a ReportBackup row and store it under the backup prefix.

    Runs inside the Celery worker (or inline as a last resort). Status/result
    are persisted on the row so the web process can poll over the shared DB.
    """
    ReportBackup = _report_backup_model()
    backup = ReportBackup.objects.filter(pk=backup_pk).first()
    if backup is None or backup.status not in (ReportBackup.STATUS_PENDING,):
        return backup
    backup.status = ReportBackup.STATUS_RUNNING
    backup.started_at = timezone.now()
    backup.save(update_fields=["status", "started_at"])
    try:
        with tempfile.TemporaryFile() as tmp:
            manifest = write_backup_zip(backup.user, tmp, window=backup.window)
            size = tmp.tell()
            tmp.seek(0)
            saved_path = default_storage.save(
                f"{get_backup_storage_prefix()}/{backup.token}.zip",
                File(tmp),
            )
        backup.file_path = saved_path
        backup.file_size = size
        backup.model_count = len(manifest["models"])
        backup.file_count = len(manifest["files"])
        backup.missing_file_count = len(manifest["missing_files"])
        backup.status = ReportBackup.STATUS_COMPLETED
        backup.completed_at = timezone.now()
        backup.error = ""
        backup.save()
        UserActivityLog = apps.get_model("microsys", "UserActivityLog")
        UserActivityLog.safe_log(
            user=backup.user,
            action="EXPORT",
            model_name="Microsys Reports Backup",
            details={
                "window": backup.window,
                "models": backup.model_count,
                "files": backup.file_count,
            },
        )
        _prune_report_backups(backup.user)
    except Exception as exc:
        logger.exception("Report backup pk=%s failed", backup_pk)
        backup.status = ReportBackup.STATUS_FAILED
        backup.completed_at = timezone.now()
        backup.error = str(exc)[:1000]
        backup.save(update_fields=["status", "completed_at", "error"])
    return backup
