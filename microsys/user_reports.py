from collections import Counter, defaultdict
from datetime import timedelta
from io import BytesIO

from django.apps import apps
from django.contrib.auth import get_user_model
from django.utils import timezone

from .translations import get_strings
from .utils import get_user_management_tier_state_for_user, translate_activity_log_model_name
from .reports import (
    build_activity_windows,
    apply_report_window,
    filter_report_eligible_activity,
    normalize_report_window,
)


def _safe_list(value):
    return list(value or []) if isinstance(value, list) else []


def _add_clean_values(target, values):
    for value in _safe_list(values):
        value = str(value or '').strip()
        if value:
            target.add(value)


def _format_duration(seconds):
    s = get_strings()
    seconds = max(0, int(seconds or 0))
    hours, remainder = divmod(seconds, 3600)
    minutes, _ = divmod(remainder, 60)
    if hours:
        return s.get('user_report_duration_hours_minutes').format(hours=hours, minutes=minutes)
    return s.get('user_report_duration_minutes').format(minutes=minutes)


def _datetime_or_none(value):
    return value if value else None


def _distinct_values(queryset, field_name):
    values = []
    for value in queryset.values_list(field_name, flat=True).distinct():
        value = str(value or '').strip()
        if value:
            values.append(value)
    return values


def build_user_report(target_user, actor=None, window='week'):
    s = get_strings()
    window = normalize_report_window(window)
    ActivityLog = apps.get_model('microsys', 'UserActivityLog')
    TrustedDevice = apps.get_model('microsys', 'TrustedDevice')
    KnownDevice = apps.get_model('microsys', 'UserKnownDevice')
    PresenceSession = apps.get_model('microsys', 'UserPresenceSession')

    activity_qs = ActivityLog._default_manager.filter(created_by=target_user).order_by('-created_at')
    eligible_activity_qs = filter_report_eligible_activity(activity_qs).order_by('-created_at')
    selected_activity_qs = apply_report_window(eligible_activity_qs, window).order_by('-created_at')
    presence_qs = PresenceSession.objects.filter(user=target_user).select_related('known_device').order_by('-last_seen_at')
    known_devices = KnownDevice.objects.filter(user=target_user).select_related('trusted_device').order_by('-last_seen_at')
    trusted_devices = TrustedDevice.objects.filter(user=target_user).order_by('-last_used_at')

    first_activity = eligible_activity_qs.order_by('created_at').first()
    last_activity = eligible_activity_qs.first()
    first_presence = presence_qs.order_by('first_seen_at').first()
    last_presence = presence_qs.first()

    def _fmt_actions(counter):
        return [
            {
                'key': key,
                'label': s.get(f'action_{str(key or "").lower()}', key or s.get('user_report_unknown')),
                'count': count,
            }
            for key, count in counter.most_common()
        ]

    def _fmt_models(counter):
        return [
            {
                'key': key,
                'label': translate_activity_log_model_name(key, strings=s),
                'count': count,
            }
            for key, count in counter.most_common()
        ]

    def _fmt_model_actions(model_action_map):
        """Merge model + action breakdowns: one entry per model, each carrying its
        own action-type breakdown and total, sorted by total descending."""
        models = []
        for model_key, action_counter in model_action_map.items():
            models.append({
                'key': model_key,
                'label': translate_activity_log_model_name(model_key, strings=s),
                'count': sum(action_counter.values()),
                'actions': _fmt_actions(action_counter),
            })
        models.sort(key=lambda m: m['count'], reverse=True)
        return models

    windows = build_activity_windows(eligible_activity_qs, strings=s)
    action_counts = Counter()
    model_counts = Counter()
    for action, model_name in selected_activity_qs.values_list('action', 'model_name'):
        action_counts[action] += 1
        model_counts[model_name or s.get('user_report_unknown')] += 1
    presence_days = set()
    total_seconds = 0
    total_requests = 0
    for item in presence_qs:
        total_seconds += int(item.estimated_seconds or 0)
        total_requests += int(item.request_count or 0)
        if item.first_seen_at:
            presence_days.add(timezone.localtime(item.first_seen_at).date())
        if item.last_seen_at:
            presence_days.add(timezone.localtime(item.last_seen_at).date())

    ip_addresses = set(_distinct_values(eligible_activity_qs, 'ip_address'))
    user_agents = set(_distinct_values(eligible_activity_qs, 'user_agent'))
    browsers = set()
    operating_systems = set()
    for device in known_devices:
        _add_clean_values(ip_addresses, device.ip_addresses)
        _add_clean_values(user_agents, device.user_agents)
        _add_clean_values(browsers, device.browser_names)
        _add_clean_values(operating_systems, device.os_names)
    for item in presence_qs:
        _add_clean_values(ip_addresses, item.ip_addresses)
        _add_clean_values(user_agents, item.user_agents)
        _add_clean_values(browsers, item.browser_names)
        _add_clean_values(operating_systems, item.os_names)

    try:
        public_registration = target_user.public_registration
    except Exception:
        public_registration = None
    try:
        profile = target_user.profile
    except Exception:
        profile = None
    tier = get_user_management_tier_state_for_user(target_user)
    generated_at = timezone.now()
    history_started_at = None
    if first_presence:
        history_started_at = first_presence.first_seen_at
    elif first_activity:
        history_started_at = first_activity.created_at

    return {
        'generated_at': generated_at,
        'selected_window': window,
        'history_started_at': history_started_at,
        'target_user': target_user,
        'profile': profile,
        'tier': tier,
        'public_registration': public_registration,
        'summary': {
            'username': target_user.get_username(),
            'display_name': target_user.get_full_name() or target_user.get_username(),
            'email': target_user.email or '',
            'is_active': bool(target_user.is_active),
            'is_staff': bool(target_user.is_staff),
            'is_superuser': bool(target_user.is_superuser),
            'date_joined': target_user.date_joined,
            'last_login': target_user.last_login,
            'activity_count': eligible_activity_qs.count(),
            'known_device_count': known_devices.count(),
            'presence_session_count': presence_qs.count(),
            'trusted_device_count': trusted_devices.filter(revoked_at__isnull=True, trusted_until__gt=generated_at).count(),
            'first_activity_at': _datetime_or_none(first_activity.created_at if first_activity else None),
            'last_activity_at': _datetime_or_none(last_activity.created_at if last_activity else None),
            'first_seen_at': _datetime_or_none(first_presence.first_seen_at if first_presence else None),
            'last_seen_at': _datetime_or_none(last_presence.last_seen_at if last_presence else None),
            'days_observed': len(presence_days),
            'estimated_seconds': total_seconds,
            'estimated_duration': _format_duration(total_seconds),
            'request_count': total_requests,
            'ip_count': len(ip_addresses),
            'browser_count': len(browsers),
            'os_count': len(operating_systems),
        },
        'action_counts': _fmt_actions(action_counts),
        'model_counts': _fmt_models(model_counts),
        'windows': windows,
        'known_devices': list(known_devices),
        'presence_sessions': list(presence_qs[:100]),
        'trusted_devices': list(trusted_devices[:100]),
        'ip_addresses': sorted(ip_addresses),
        'user_agents': sorted(user_agents),
        'browsers': sorted(browsers),
        'operating_systems': sorted(operating_systems),
        'recent_activity': list(eligible_activity_qs[:25]),
        'activity_qs': eligible_activity_qs,
        'selected_activity_qs': selected_activity_qs,
    }


def build_user_report_xlsx(report, window=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    s = get_strings()
    wb = Workbook()
    summary = wb.active
    summary.title = s.get('user_report_sheet_summary')[:31]

    def clean_cell_value(value):
        if isinstance(value, timezone.datetime):
            if timezone.is_aware(value):
                value = timezone.localtime(value)
            return value.replace(tzinfo=None)
        return value

    def write_rows(ws, rows):
        for row in rows:
            ws.append([clean_cell_value(value) for value in row])
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for column in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in column)
            ws.column_dimensions[column[0].column_letter].width = min(max(max_len + 2, 12), 60)

    summary_rows = [
        [s.get('user_report_field'), s.get('user_report_value')],
    ]
    for key in (
        'username',
        'display_name',
        'email',
        'is_active',
        'is_staff',
        'is_superuser',
        'date_joined',
        'last_login',
        'activity_count',
        'known_device_count',
        'presence_session_count',
        'trusted_device_count',
        'estimated_duration',
        'request_count',
        'ip_count',
    ):
        summary_rows.append([s.get(f'user_report_{key}'), report['summary'].get(key)])
    write_rows(summary, summary_rows)

    ws = wb.create_sheet(s.get('user_report_sheet_activity')[:31])
    rows = [[s.get('user_report_action'), s.get('user_report_count')]]
    rows.extend([[item['label'], item['count']] for item in report['action_counts']])
    write_rows(ws, rows)

    ws = wb.create_sheet(s.get('user_report_sheet_presence')[:31])
    rows = [[
        s.get('user_report_device'),
        s.get('user_report_first_seen'),
        s.get('user_report_last_seen'),
        s.get('user_report_estimated_time'),
        s.get('user_report_requests'),
        s.get('user_report_ip_addresses'),
    ]]
    for item in report['presence_sessions']:
        rows.append([
            item.device_label,
            item.first_seen_at,
            item.last_seen_at,
            _format_duration(item.estimated_seconds),
            item.request_count,
            ', '.join(_safe_list(item.ip_addresses)),
        ])
    write_rows(ws, rows)

    ws = wb.create_sheet(s.get('user_report_sheet_devices')[:31])
    rows = [[
        s.get('user_report_device'),
        s.get('trusted_device_badge'),
        s.get('user_report_first_seen'),
        s.get('user_report_last_seen'),
        s.get('user_report_ip_addresses'),
        s.get('user_report_browsers'),
        s.get('user_report_operating_systems'),
    ]]
    for item in report['known_devices']:
        rows.append([
            item.device_label,
            bool(item.trusted_device_id and item.trusted_device and item.trusted_device.is_active),
            item.first_seen_at,
            item.last_seen_at,
            ', '.join(_safe_list(item.ip_addresses)),
            ', '.join(_safe_list(item.browser_names)),
            ', '.join(_safe_list(item.os_names)),
        ])
    write_rows(ws, rows)

    ws = wb.create_sheet(s.get('user_report_sheet_logs')[:31])
    rows = [[
        s.get('user_report_timestamp'),
        s.get('user_report_action'),
        s.get('user_report_model'),
        s.get('user_report_ip_address'),
        s.get('user_report_user_agent'),
    ]]
    selected_qs = report.get('selected_activity_qs') or report['activity_qs']
    if window:
        selected_qs = apply_report_window(selected_qs, normalize_report_window(window)).order_by('-created_at')
    for item in selected_qs[:1000]:
        rows.append([
            item.created_at,
            s.get(f'action_{str(item.action or "").lower()}', item.action),
            translate_activity_log_model_name(item.model_name, strings=s),
            item.ip_address,
            item.user_agent,
        ])
    write_rows(ws, rows)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream.getvalue()
